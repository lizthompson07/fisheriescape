import os
import datetime

import xlsxwriter
from openpyxl import load_workbook

import projects2.api.views as ApiViews
import projects2.test.FactoryFloor as factory

from django.test import tag, RequestFactory
from django.utils import timezone
from projects2 import models as models, reports
from shared_models import models as shared_models
from shared_models.test.common_tests import CommonTest

fixtures_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'fixtures')
standard_fixtures = [file for file in os.listdir(fixtures_dir)]

req_factory = RequestFactory()


def create_projects(count, year):
    projects = factory.ProjectFactory.create_batch(count)
    for y in year:
        for p in projects:
            fy_list = shared_models.FiscalYear.objects.filter(pk__in=year)
            p.fiscal_years.set(fy_list)
            p.save()
            factory.ProjectYearFactory.create(project=p, start_date=datetime.datetime(year=y, month=1, day=1, tzinfo=timezone.get_current_timezone()))


@tag("ExportProjectSummaryReport", "Report")
class ExportProjectSummaryReportTest(CommonTest):
    fixtures = standard_fixtures

    EXPECTED_HEADER_ROWS = 2

    EXPECTED_N_YEAR_2020 = 10
    EXPECTED_N_YEAR_2021 = 5
    EXPECTED_N_YEAR_2020_2021 = 20

    EXPECTED_TOTAL_PROJECTS = EXPECTED_N_YEAR_2020 + EXPECTED_N_YEAR_2021 + EXPECTED_N_YEAR_2020_2021

    EXPECTED_START_DATE_COLUMN = 'I'
    EXPECTED_END_DATE_COLUMN = 'J'

    def setUp(self):
        create_projects(self.EXPECTED_N_YEAR_2020, [2020])
        create_projects(self.EXPECTED_N_YEAR_2021, [2021])
        create_projects(self.EXPECTED_N_YEAR_2020_2021, [2020, 2021])

    # get_project_year_queryset is actually part of the API view, but the Export Project Summary Relies on it
    def test_get_project_year_queryset(self):

        user = self.get_and_login_user(in_group="projects_admin")

        request = req_factory.get("")
        request.user = user
        # simulate a GET request
        request.GET = {"fiscal_year": 2020}

        qs = ApiViews.get_project_year_queryset(request=request)

        # querying for projects that are in 2020
        self.assertEqual(len(qs), (self.EXPECTED_N_YEAR_2020 + self.EXPECTED_N_YEAR_2020_2021))

    def test_export_project_summary_query_year(self):
        user = self.get_and_login_user(in_group="projects_admin")

        request = req_factory.get("")
        request.user = user
        # simulate a GET request
        request.GET = {"fiscal_year": 2020}

        response = reports.export_project_summary(request)

        wb = load_workbook(response)
        # should be only two worksheets, the one for 2020 and the template
        self.assertEqual(len(wb.sheetnames), 2)
        self.assertEqual(wb.get_sheet_by_name("2019-2020").max_row,
                         (self.EXPECTED_N_YEAR_2020 + self.EXPECTED_N_YEAR_2020_2021 + self.EXPECTED_HEADER_ROWS))

    def test_export_project_summary_query_empty(self):
        user = self.get_and_login_user(in_group="projects_admin")

        request = req_factory.get("")
        request.user = user
        # simulate a GET request
        request.GET = {}

        response = reports.export_project_summary(request)

        wb = load_workbook(response)
        # should be three worksheets, 2020, 2021 and the template
        self.assertEqual(3, len(wb.sheetnames))

        self.assertEqual(wb.get_sheet_by_name("2019-2020").max_row,
                         (self.EXPECTED_N_YEAR_2020 + self.EXPECTED_N_YEAR_2020_2021 + self.EXPECTED_HEADER_ROWS))

        self.assertEqual(wb.get_sheet_by_name("2020-2021").max_row,
                         (self.EXPECTED_N_YEAR_2021 + self.EXPECTED_N_YEAR_2020_2021 + self.EXPECTED_HEADER_ROWS))

    def test_date_format_export_project_summary_query_empty(self):
        user = self.get_and_login_user(in_group="projects_admin")

        request = req_factory.get("")
        request.user = user
        # simulate a GET request
        request.GET = {}

        response = reports.export_project_summary(request)

        wb = load_workbook(response)

        # should be three worksheets, 2020, 2021 and the template
        self.assertEqual(3, len(wb.sheetnames))
        sheet = wb.get_sheet_by_name("2019-2020")
        self.assertEqual(sheet[self.EXPECTED_START_DATE_COLUMN + str(3)].value, "2020-01-01")
