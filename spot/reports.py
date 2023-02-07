from django.http import HttpResponse
from . import models
from . import filters
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.utils.cell import get_column_letter
from django.core.exceptions import ObjectDoesNotExist

green = 'C4D79B'


def export_data(request):
    data = models.Data.objects.all()
    data_filter = filters.DataFilter(request.GET, queryset=data).qs
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-Data.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Data'
    columns = [
        'Agreement Number',
        'Project Number',
        'Project Name',
        'Funding Year',
        'Funding Source',
        'Project Authority',
        'Biological Process Type - Level 1',
        'Biological Process Type - Level 2',
        'Biological Process Type - level 3',
        'Activity Type - level 1',
        'Activity Type - level 2',
        'Activity Type - level 3',
        'Area',
        'Location/River',
        'Species',
        'Sub District Area',
        'Stock Management Area',
        'CU Index',
        'DU',
        'DU Number',
        'CU Name',
        'Watershed',
        'First Nations',
        'Funding Recipient',
        'Primary Organization or Program',
        'Hatchery',
        'Samples Collected',
        'Samples Collected Comment',
        'Sample Collected Database',
        'Shared Drive',
        'Was sample collection data entered into database(s)?',
        'Was sample data quality check complete?',
        'Sampling Entry Format',
        'Data Product(s)',
        'Data Products Database(s)',
        'Data Products Comment',
        'date_last_modified',
        'last_modified_by',
    ]

    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        worksheet.column_dimensions[get_column_letter(col_num)].width = 15
        cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.value = column_title

    for obj in data_filter:
        proj = models.Project.objects.get(pk=obj.project.id)
        if obj.river and obj.river.cu_name:
            cu_tmp = obj.river.cu_name.name
        else:
            cu_tmp = None
        if obj.river and obj.river.watershed:
            wtr_tmp = obj.river.watershed.name
        else:
            wtr_tmp = None
        if obj.river and obj.river.sub_district_area:
            sub_tmp = obj.river.sub_district_area.name
        else:
            sub_tmp = None
        if obj.river and obj.river.cu_index:
            cui_tmp = obj.river.cu_index.name
        else:
            cui_tmp = None
        if obj.river and obj.river.du:
            du_tmp = obj.river.du.name
        else:
            du_tmp = None
        funding_year = models.FundingYears.objects.filter(project=obj.project.id)
        row_num += 1
        row = [
            proj.agreement_number,
            proj.project_number,
            proj.name,
            ", ".join(i.funding_year for i in funding_year.all()),
            ", ".join(i.name for i in proj.funding_sources.all()),
            ", ".join(i.full_name for i in proj.DFO_project_authority.all()),
            proj.biological_process_type_1,
            proj.biological_process_type_2,
            ", ".join(i.name for i in proj.biological_process_type_3.all()),
            ", ".join(i.name for i in proj.activity_type_1.all()),
            ", ".join(i.name for i in proj.activity_type_2.all()),
            ", ".join(i.name for i in proj.activity_type_3.all()),
            proj.area,
            obj.river.name if obj.river else None,
            obj.river.species if obj.river else None,
            sub_tmp,
            obj.river.stock_management_unit if obj.river else None,
            cui_tmp,
            du_tmp,
            obj.river.du_number if obj.river else None,
            cu_tmp,
            wtr_tmp,
            proj.first_nation.name if proj.first_nation else None,
            proj.funding_recipient,
            proj.organization_program.name if proj.organization_program else None,
            ", ".join(i.name for i in proj.hatchery_name.all()),
            obj.samples_collected,
            obj.samples_collected_comment,
            obj.samples_collected_database,
            obj.shared_drive,
            obj.sample_entered_database,
            obj.data_quality_check,
            obj.sample_format,
            obj.data_products,
            obj.data_products_database,
            obj.data_products_comment,
            obj.date_last_modified,
            obj.last_modified_by.get_full_name() if obj.last_modified_by else None,
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def export_method(request):
    method = models.Method.objects.all()
    method_filter = filters.MethodFilter(request.GET, queryset=method).qs
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-Method.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Method'
    columns = [
        'Agreement Number',
        'Unique Method Number',
        'Project Number',
        'Project Name',
        'Funding Year',
        'Funding Source',
        'Project Authority',
        'Biological Process Type - Level 1',
        'Biological Process Type - level 2',
        'Biological Process Type - level 3',
        'Activity Type - level 1',
        'Activity Type - level 2',
        'Activity Type - level 3',
        'Area',
        'Location/River',
        'Species',
        'Sub District Area',
        'Stock Management Unit',
        'CU Index',
        'DU',
        'DU Number',
        'CU Name',
        'Watershed',
        'First Nations',
        'Funding Recipient',
        'Primary Organization or Program',
        'Hatchery',
        'Field Work Methods',
        'Planning Method Type',
        'Sample Processing Method Type',
        'Scale Processing Location',
        'Otolith Processing Location',
        'DNA Processing Location',
        'Heads Processing Location',
        'Instrument Data Processing Location',
        'date_last_modified',
        'last_modified_by',
    ]

    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = 15
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.value = column_title

    method_document_worksheet = workbook.create_sheet("Method Document", 1)
    method_document_columns = [
        'Agreement Number',
        'Unique Method Number',
        'Project Number',
        'Project Name',
        'Project Authority',
        'Biological Process Type - Level 1',
        'Biological Process Type - level 2',
        'Biological Process Type - level 3',
        'Activity Type - level 1',
        'Activity Type - level 2',
        'Activity Type - level 3',
        'Area',
        'Location/River',
        'Species',
        'Sub District Area',
        'Stock Management Unit',
        'CU index',
        'DU',
        'DU Number',
        'CU Name',
        'Watershed',
        'First Nations',
        'Funding Recipient',
        'Primary Organization or Program'
        'Hatchery',
        'Method Document Type',
        'Author',
        'Year of Publication',
        'Title',
        'Reference Number',
        'Document Link',
        'date_last_modified',
        'last_modified_by',
    ]
    row_num = 1

    for col_num, column_title in enumerate(method_document_columns, 1):
        method_document_worksheet.column_dimensions[get_column_letter(col_num)].width = 15
        cell = method_document_worksheet.cell(row=row_num, column=col_num)
        cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.value = column_title

    meth_num = 1
    for obj in method_filter:
        proj = models.Project.objects.get(pk=obj.project.id)
        all_cu_tmp = ""
        all_wtr_tmp = ""
        all_du_tmp = ""
        all_cui_tmp = ""
        all_sub_tmp = ""
        for i in proj.river.all():
            if i.cu_name:
                all_cu_tmp += str(i.cu_name.name + ", ")
            if i.watershed:
                all_wtr_tmp += str(i.watershed.name + ", ")
            if i.cu_index:
                all_cui_tmp += str(i.cu_index.name + ", ")
            if i.du:
                all_du_tmp += str(i.du.name + ", ")
            if i.sub_district_area:
                all_sub_tmp += str(i.sub_district_area.name + ", ")
        funding_year = models.FundingYears.objects.filter(project=obj.project.id)
        row_num += 1
        row = [
            proj.agreement_number,
            obj.unique_method_number,
            proj.project_number,
            proj.name,
            ", ".join(i.funding_year for i in funding_year.all()),
            ", ".join(i.name for i in proj.funding_sources.all()),
            ", ".join(i.full_name for i in proj.DFO_project_authority.all()),
            proj.biological_process_type_1,
            proj.biological_process_type_2,
            ", ".join(i.name for i in proj.biological_process_type_3.all()),
            ", ".join(i.name for i in proj.activity_type_1.all()),
            ", ".join(i.name for i in proj.activity_type_2.all()),
            ", ".join(i.name for i in proj.activity_type_3.all()),
            proj.area,
            ", ".join(i.name for i in proj.river.all()),
            ", ".join(i.species for i in proj.river.all()),
            all_sub_tmp,
            ", ".join(i.stock_management_unit if i.stock_management_unit else "" for i in proj.river.all()),
            all_cui_tmp,
            all_du_tmp,
            ", ".join(i.du_number if i.du_number else "" for i in proj.river.all()),
            all_cu_tmp,
            all_wtr_tmp,
            proj.first_nation.name if proj.first_nation else None,
            proj.funding_recipient,
            proj.organization_program.name if proj.organization_program else None,
            ", ".join(i.name for i in proj.hatchery_name.all()),
            obj.field_work_method_type,
            obj.planning_method_type,
            obj.sample_processing_method_type,
            obj.scale_processing_location,
            obj.otolith_processing_location,
            obj.DNA_processing_location,
            obj.heads_processing_location,
            obj.instrument_data_processing_location,
            obj.date_last_modified,
            obj.last_modified_by.get_full_name() if obj.last_modified_by else None,
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

        method_documents = models.MethodDocument.objects.filter(method=obj.id)
        for method in method_documents:
            meth_num += 1
            method_rows = [
                proj.agreement_number,
                method.unique_method_number,
                proj.project_number,
                proj.name,
                ", ".join(i.full_name for i in proj.DFO_project_authority.all()),
                proj.biological_process_type_1,
                proj.biological_process_type_2,
                ", ".join(i.name for i in proj.biological_process_type_3.all()),
                ", ".join(i.name for i in proj.activity_type_1.all()),
                ", ".join(i.name for i in proj.activity_type_2.all()),
                ", ".join(i.name for i in proj.activity_type_3.all()),
                proj.area,
                ", ".join(i.name for i in proj.river.all()),
                ", ".join(i.species for i in proj.river.all()),
                all_sub_tmp,
                ", ".join(i.stock_management_unit if i.stock_management_unit else "" for i in proj.river.all()),
                all_cui_tmp,
                all_du_tmp,
                ", ".join(i.du_number if i.du_number else "" for i in proj.river.all()),
                all_wtr_tmp,
                all_cu_tmp,
                all_wtr_tmp,
                proj.first_nation.name if proj.first_nation else None,
                proj.funding_recipient,
                proj.organization_program.name if proj.organization_program else None,
                ", ".join(i.name for i in proj.hatchery_name.all()),
                method.method_document_type,
                method.authors,
                method.publication_year,
                method.title,
                method.reference_number,
                method.document_link,
                method.date_last_modified,
                method.last_modified_by.get_full_name() if method.last_modified_by else None,
            ]
            for col_num, cell_value in enumerate(method_rows, 1):
                cell = method_document_worksheet.cell(row=meth_num, column=col_num)
                cell.value = cell_value

    workbook.save(response)
    return response


def export_reports(request):
    reports = models.Reports.objects.all()
    reports_filter = filters.ReportsFilter(request.GET, queryset=reports).qs
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-Reports.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Reports'
    columns = [
        'Agreement Number',
        'Project Number',
        'Project Name',
        'Funding Year',
        'Funding Source',
        'Project Authority',
        'Biological Process Type - Level 1',
        'Biological Process Type - level 2',
        'Biological Process Type - level 3',
        'Activity Type - level 1',
        'Activity Type - level 2',
        'Activity Type - level 3',
        'Area',
        'Location/River',
        'Species',
        'Sub District Area',
        'Stock Management Unit',
        'CU index',
        'DU',
        'DU Number',
        'CU Name',
        'Watershed',
        'First Nations',
        'Funding Recipient',
        'Primary Organization or Program ',
        'Hatchery',
        'Report Timeline',
        'Report Type',
        'Document Name',
        'Document Author',
        'Document Reference Information',
        'Document Link',
        'Was this report Published?',
        'date_last_modified',
        'last_modified_by',

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = 15
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.value = column_title

    for obj in reports_filter:
        proj = models.Project.objects.get(pk=obj.project.id)
        all_cu_tmp = ""
        all_wtr_tmp = ""
        all_du_tmp = ""
        all_cui_tmp = ""
        all_sub_tmp = ""
        for i in proj.river.all():
            if i.cu_name:
                all_cu_tmp += str(i.cu_name.name + ", ")
            if i.watershed:
                all_wtr_tmp += str(i.watershed.name + ", ")
            if i.cu_index:
                all_cui_tmp += str(i.cu_index.name + ", ")
            if i.du:
                all_du_tmp += str(i.du.name + ", ")
            if i.sub_district_area:
                all_sub_tmp += str(i.sub_district_area.name + ", ")
        funding_year = models.FundingYears.objects.filter(project=obj.project.id)
        row_num += 1
        row = [
            proj.agreement_number,
            proj.project_number,
            proj.name,
            ", ".join(i.funding_year for i in funding_year.all()),
            ", ".join(i.name for i in proj.funding_sources.all()),
            ", ".join(i.full_name for i in proj.DFO_project_authority.all()),
            proj.biological_process_type_1,
            proj.biological_process_type_2,
            ", ".join(i.name for i in proj.biological_process_type_3.all()),
            ", ".join(i.name for i in proj.activity_type_1.all()),
            ", ".join(i.name for i in proj.activity_type_2.all()),
            ", ".join(i.name for i in proj.activity_type_3.all()),
            proj.area,
            ", ".join(i.name for i in proj.river.all()),
            ", ".join(i.species for i in proj.river.all()),
            all_sub_tmp,
            ", ".join(i.stock_management_unit if i.stock_management_unit else "" for i in proj.river.all()),
            all_cui_tmp,
            all_du_tmp,
            ", ".join(i.du_number if i.du_number else "" for i in proj.river.all()),
            all_cu_tmp,
            all_wtr_tmp,
            proj.first_nation.name if proj.first_nation else None,
            proj.funding_recipient,
            proj.organization_program.name if proj.organization_program else None,
            ", ".join(i.name for i in proj.hatchery_name.all()),
            obj.report_timeline,
            obj.report_type,
            obj.document_name,
            obj.document_author,
            obj.document_reference_information,
            obj.document_link,
            obj.published,
            obj.date_last_modified,
            obj.last_modified_by.get_full_name() if obj.last_modified_by else None,
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response


def export_project_full(request):
    project = models.Project.objects.all()
    project_filter = filters.ProjectFilter(request.GET, queryset=project).qs
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-Full-Projects.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Projects'

    ########################
    # PROJECT SHEET/HEADER #
    columns = [
        'Agreement Number',
        'Project Number',

        'Agreement History',
        'Project Name',
        'Project Description',
        'Funding Year',
        'Start Date',
        'End Date',

        'Area',
        'Other Species',
        'Ecosystem Type',
        'Lake System(s)',
        'Watershed(s)',
        'Other Site Info',

        'Life Stage',
        'Aquaculture License Number',
        'Water License Number',
        'Hatchery Name',
        'DFO Tenure',

        'Project Stage',
        'Biological Process Type - Level 1',
        'Activity Type - level 1',
        'Monitoring Approach',
        'Biological Process Type - level 3',
        'Biological Process Type - level 2',
        'Activity Type - level 2',
        'Activity Type - level 3',
        'Project Purpose',
        'Category Comments',

        'Link to other DFO Programs',
        'Linked DFO Program Project Reference',
        'Link to other Government Departments',
        'Policy Connections',

        'DFO Project Authority',
        'DFO Area Chief',
        'DFO IAA',
        'DFO Resource Manager',
        'Funding Recipient',
        'First Nation/Tribal Council',
        'Contact',
        'Contact Role',
        'DFO Technicians/Biologists',
        'Contractors',
        'Contractor Contact',
        'Partner',
        'Partner Contact',

        'Agreement Database',
        'Agreement Comment',
        'Funding Sources',
        'Other Funding Sources',
        'Agreement Type',
        'Primary Organization or Program',

        'date_last_modified',
        'last_modified_by',


    ]
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = 15
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.value = column_title

    ##############################
    # FUNDING YEARS SHEET/HEADER #
    funding_worksheet = workbook.create_sheet("Funding", 1)
    funding_columns = [
        'Agreement Number',
        'Project Number',
        'Project Name',
        'Funding Year',
        'Funding Source',
        'Project Authority',
        'First Nation',
        'Location / River',
        'Species',
        'Sub District Area',
        'Stock Management Unit',
        'CU Index',
        'CU Name',
        'DU',
        'DU Number',
        'Watershed',
        'Funding Recipient',
        'Biological Process Type - Level 1',
        'Biological Process Type - Level 2',
        'Biological Process Type - Level 3',
        'Activity Type - level 1',
        'Activity Type - level 2',
        'Activity Type - level 3',
        'Area',
        'Agreement Cost',
        'Project Cost',
        'date_last_modified',
        'last_modified_by',
    ]
    row_num = 1
    for col_num, column_title in enumerate(funding_columns, 1):
        funding_worksheet.column_dimensions[get_column_letter(col_num)].width = 15
        cell = funding_worksheet.cell(row=row_num, column=col_num)
        cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.value = column_title

    ##############################
    # Rivers SHEET/HEADER #
    river_worksheet = workbook.create_sheet("Rivers", 9)
    river_columns = [
        'Agreement Number',
        'Project Number',
        'Project Name',
        'Project Authority',
        'Biological Process Type - Level 1',
        'Biological Process Type - level 2',
        'Biological Process Type - level 3',
        'Activity Type - level 1',
        'Activity Type - level 2',
        'Activity Type - level 3',
        'Area',
        'First Nations',
        'Funding Recipient',
        'Primary Organization or Program',
        'Hatchery',
        'Name',
        'Latitude',
        'Longitude',
        'Sub District Area',
        'Species',
        'Stock Management Unit',
        'CU index',
        'CU Name',
        'Pop ID',
        'DU',
        'DU Number',
        'Watershed',
    ]
    row_num = 1
    for col_num, column_title in enumerate(river_columns, 1):
        river_worksheet.column_dimensions[get_column_letter(col_num)].width = 15
        cell = river_worksheet.cell(row=row_num, column=col_num)
        cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.value = column_title

    ############################
    # PROJECT CERTIFIED HEADER #
    certified_worksheet = workbook.create_sheet("Project Certified", 10)
    certified_columns = [
        'Agreement Number',
        'Project Number',
        'Project Name',
        'Biological Process Type - Level 1',
        'Biological Process Type - level 2',
        'Biological Process Type - level 3',
        'Activity Type - level 1',
        'Activity Type - level 2',
        'Activity Type - level 3',
        'Area',
        'Location/River',
        'Species',
        'Sub District Area',
        'Stock Management Unit',
        'CU index',
        'DU',
        'DU Number',
        'CU Name',
        'Watershed',
        'Certified Date',
        'Certified By',
    ]
    row_num = 1
    for col_num, column_title in enumerate(certified_columns, 1):
        certified_worksheet.column_dimensions[get_column_letter(col_num)].width = 15
        cell = certified_worksheet.cell(row=row_num, column=col_num)
        cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.value = column_title


    ##########################
    # ACTIVITY OUTCOME SHEET/HEADER #
    activity_outcome_worksheet = workbook.create_sheet("Activities & Outcomes", 2)
    activity_outcome_columns = [
        'Agreement Number',
        'Unique Activity Outcome Number',
        'Project Number',
        'Project Name',
        'Funding Year',
        'Funding Source',
        'Project Authority',
        'Biological Process Type - Level 1',
        'Biological Process Type - level 2',
        'Biological Process Type - level 3',
        'Activity Type - level 1',
        'Activity Type - level 2',
        'Activity Type - level 3',
        'Area',
        'Location/River',
        'Species',
        'Sub District Area',
        'Stock Management Unit',
        'CU index',
        'DU',
        'DU Number',
        'CU Name',
        'Watershed',
        'First Nations',
        'Funding Recipient',
        'Primary Organization or Program',
        'Hatchery',
        'Task Description',
        'Element Title',
        'Activity Title',
        'PST Requirement Identified?',
        'Activity & Outcome Category',
        'SIL Requirement',
        'Expected Result(s)',
        'Products/Reports to Provide DFO',
        'Outcome Comment',
        'date_last_modified',
        'last_modified_by',

    ]
    row_num = 1
    for col_num, column_title in enumerate(activity_outcome_columns, 1):
        activity_outcome_worksheet.column_dimensions[get_column_letter(col_num)].width = 15
        cell = activity_outcome_worksheet.cell(row=row_num, column=col_num)
        cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.value = column_title

    #########################
    # SAMPLE OUTCOME HEADER #
    sample_outcome_worksheet = workbook.create_sheet("Sample Outcomes", 3)
    sample_outcomes_columns = [
        'Agreement Number',
        'Project Number',
        'Project Name',
        'Funding Year',
        'Funding Source',
        'Project Authority',
        'Biological Process Type - Level 1',
        'Biological Process Type - level 2',
        'Biological Process Type - level 3',
        'Activity Type - level 1',
        'Activity Type - level 2',
        'Activity Type - level 3',
        'Area',
        'Location/River',
        'Species',
        'Sub District Area',
        'Stock Management Unit',
        'CU index',
        'DU',
        'DU Number',
        'CU Name',
        'Watershed',
        'First Nations',
        'Funding Recipient',
        'Primary Organization or Program',
        'Hatchery',
        'Unique Activity and Outcome Number',
        'Sampling Outcome',
        'Were outcomes delivered?',
        'Quality of Outcome',
        'Sample Outcome Comment',
        'date_last_modified',
        'last_modified_by',
    ]
    row_num = 1

    for col_num, column_title in enumerate(sample_outcomes_columns, 1):
        sample_outcome_worksheet.column_dimensions[get_column_letter(col_num)].width = 15
        cell = sample_outcome_worksheet.cell(row=row_num, column=col_num)
        cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.value = column_title

    ##################################
    # REPORTING OUTCOME SHEET/HEADER #
    reporting_outcome_worksheet = workbook.create_sheet("Report Outcomes", 4)
    report_outcomes_columns = [
        'Agreement Number',
        'Project Number',
        'Project Name',
        'Funding Year',
        'Funding Source',
        'Project Authority',
        'Biological Process Type - Level 1',
        'Biological Process Type - level 2',
        'Biological Process Type - level 3',
        'Activity Type - level 1',
        'Activity Type - level 2',
        'Activity Type - level 3',
        'Area',
        'Location/River',
        'Species',
        'Sub District Area',
        'Stock Management Unit',
        'CU index',
        'DU',
        'DU Number',
        'CU Name',
        'Watershed',
        'First Nations',
        'Funding Recipient',
        'Primary Organization or Program',
        'Hatchery',
        'Unique Activity and Outcome Number',
        'Reporting Outcome',
        'Task ID',
        'Task Name',
        'Task Description',
        'Task Met',
        'Site',
        'Report Link',
        'Reporting Outcome Comment',
        'Reporting Outcome Metric',
        'Reporting Outcome Metric Unit',
        'Data Sharing',
        'date_last_modified',
        'last_modified_by',
    ]
    row_num = 1

    for col_num, column_title in enumerate(report_outcomes_columns, 1):
        reporting_outcome_worksheet.column_dimensions[get_column_letter(col_num)].width = 15
        cell = reporting_outcome_worksheet.cell(row=row_num, column=col_num)
        cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.value = column_title

    #####################
    # DATA SHEET/HEADER #
    data_worksheet = workbook.create_sheet("Data", 5)
    data_columns = [
        'Agreement Number',
        'Project Number',
        'Project Name',
        'Funding Year',
        'Funding Source',
        'Project Authority',
        'Biological Process Type - Level 1',
        'Biological Process Type - Level 2',
        'Biological Process Type - level 3',
        'Activity Type - level 1',
        'Activity Type - level 2',
        'Activity Type - level 3',
        'Area',
        'Location/River',
        'Species',
        'Sub District Area',
        'Stock Management Area',
        'CU Index',
        'DU',
        'DU Number',
        'CU Name',
        'Watershed',
        'First Nations',
        'Funding Recipient',
        'Primary Organization or Program',
        'Hatchery',
        'Samples Collected',
        'Samples Collected Comment',
        'Sample Collected Database',
        'Shared Drive',
        'Was sample collection data entered into database(s)?',
        'Was sample data quality check complete?',
        'Sampling Entry Format',
        'Data Product(s)',
        'Data Products Database(s)',
        'Data Products Comment',
        'date_last_modified',
        'last_modified_by',
    ]

    row_num = 1

    for col_num, column_title in enumerate(data_columns, 1):
        cell = data_worksheet.cell(row=row_num, column=col_num)
        data_worksheet.column_dimensions[get_column_letter(col_num)].width = 15
        cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.value = column_title

    #######################
    # METHOD SHEET/HEADER #
    method_worksheet = workbook.create_sheet("Method", 6)
    method_columns = [
        'Agreement Number',
        'Unique Method Number',
        'Project Number',
        'Project Name',
        'Funding Year',
        'Funding Source',
        'Project Authority',
        'Biological Process Type - Level 1',
        'Biological Process Type - level 2',
        'Biological Process Type - level 3',
        'Activity Type - level 1',
        'Activity Type - level 2',
        'Activity Type - level 3',
        'Area',
        'Location/River',
        'Species',
        'Sub District Area',
        'Stock Management Unit',
        'CU Index',
        'DU',
        'DU Number',
        'CU Name',
        'Watershed',
        'First Nations',
        'Funding Recipient',
        'Primary Organization or Program',
        'Hatchery',
        'Field Work Methods',
        'Planning Method Type',
        'Sample Processing Method Type',
        'Scale Processing Location',
        'Otolith Processing Location',
        'DNA Processing Location',
        'Heads Processing Location',
        'Instrument Data Processing Location',
        'date_last_modified',
        'last_modified_by',
    ]

    row_num = 1

    for col_num, column_title in enumerate(method_columns, 1):
        method_worksheet.column_dimensions[get_column_letter(col_num)].width = 15
        cell = method_worksheet.cell(row=row_num, column=col_num)
        cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.value = column_title

    ################################
    # METHOD DOCUMENT SHEET/HEADER #
    method_document_worksheet = workbook.create_sheet("Method Document", 7)
    method_document_columns = [
        'Agreement Number',
        'Unique Method Number',
        'Project Number',
        'Project Name',
        'Project Authority',
        'Biological Process Type - Level 1',
        'Biological Process Type - level 2',
        'Biological Process Type - level 3',
        'Activity Type - level 1',
        'Activity Type - level 2',
        'Activity Type - level 3',
        'Area',
        'Location/River',
        'Species',
        'Sub District Area',
        'Stock Management Unit',
        'CU index',
        'DU',
        'DU Number',
        'CU Name',
        'Watershed',
        'First Nations',
        'Funding Recipient',
        'Primary Organization or Program'
        'Hatchery',
        'Method Document Type',
        'Author',
        'Year of Publication',
        'Title',
        'Reference Number',
        'Document Link',
        'date_last_modified',
        'last_modified_by',
    ]
    row_num = 1

    for col_num, column_title in enumerate(method_document_columns, 1):
        method_document_worksheet.column_dimensions[get_column_letter(col_num)].width = 15
        cell = method_document_worksheet.cell(row=row_num, column=col_num)
        cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.value = column_title

    #################
    # REPORT SHEET/HEADER #
    report_worksheet = workbook.create_sheet("Reports", 8)
    report_columns = [
        'Agreement Number',
        'Project Number',
        'Project Name',
        'Funding Year',
        'Funding Source',
        'Project Authority',
        'Biological Process Type - Level 1',
        'Biological Process Type - level 2',
        'Biological Process Type - level 3',
        'Activity Type - level 1',
        'Activity Type - level 2',
        'Activity Type - level 3',
        'Area',
        'Location/River',
        'Species',
        'Sub District Area',
        'Stock Management Unit',
        'CU index',
        'DU',
        'DU Number',
        'CU Name',
        'Watershed',
        'First Nations',
        'Funding Recipient',
        'Primary Organization or Program ',
        'Hatchery',
        'Report Timeline',
        'Report Type',
        'Document Name',
        'Document Author',
        'Document Reference Information',
        'Document Link',
        'Was this report Published?',
        'date_last_modified',
        'last_modified_by',

    ]
    row_num = 1

    for col_num, column_title in enumerate(report_columns, 1):
        report_worksheet.column_dimensions[get_column_letter(col_num)].width = 15
        cell = report_worksheet.cell(row=row_num, column=col_num)
        cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.value = column_title

    ###########
    # PROJECT #
    repo_num = 1
    sampo_num = 1
    row_num = 1
    meth_num = 1
    methd_num = 1
    river_num = 1
    cert_num = 1
    fund_num = 1
    obj_num = 1
    data_num = 1
    rep_num = 1

    for obj in project_filter:
        row_num += 1
        all_cu_tmp = ""
        all_wtr_tmp = ""
        all_sub_tmp = ""
        all_cui_tmp = ""
        all_du_tmp = ""
        for i in obj.river.all():
            if i.cu_name:
                all_cu_tmp += str(i.cu_name.name + ", ")
            if i.watershed:
                all_wtr_tmp += str(i.watershed.name + ", ")
            if i.sub_district_area:
                all_sub_tmp += str(i.sub_district_area.name + ", ")
            if i.cu_index:
                all_cui_tmp += str(i.cu_index.name + ", ")
            if i.du:
                all_du_tmp += str(i.du.name + ", ")
        funding_year = models.FundingYears.objects.filter(project=obj.id)
        row = [
            obj.agreement_number,
            obj.project_number,

            ", ".join(i.name for i in obj.agreement_history.all()),
            obj.name,
            obj.project_description,
            ", ".join(i.funding_year for i in funding_year),
            obj.start_date,
            obj.end_date,

            obj.area,
            obj.other_species,
            ", ".join(i.name for i in obj.ecosystem_type.all()),
            ", ".join(i.name for i in obj.lake_system.all()),
            all_wtr_tmp,
            obj.other_site_info,

            ", ".join(i.name for i in obj.salmon_life_stage.all()),
            obj.aquaculture_license_number,
            obj.water_license_number,
            ", ".join(i.name for i in obj.hatchery_name.all()),
            obj.DFO_tenure,

            obj.project_stage,
            obj.biological_process_type_1,
            ", ".join(i.name for i in obj.activity_type_1.all()),
            ", ".join(i.name for i in obj.monitoring_approach.all()),
            ", ".join(i.name for i in obj.biological_process_type_3.all()),
            obj.biological_process_type_2,
            ", ".join(i.name for i in obj.activity_type_2.all()),
            ", ".join(i.name for i in obj.activity_type_3.all()),
            ", ".join(i.name for i in obj.project_purpose.all()),
            obj.category_comments,

            obj.DFO_link,
            obj.DFO_program_reference,
            obj.government_organization,
            obj.policy_connection,

            ", ".join(i.full_name for i in obj.DFO_project_authority.all()),
            ", ".join(i.full_name for i in obj.DFO_area_chief.all()),
            ", ".join(i.full_name for i in obj.DFO_IAA.all()),
            ", ".join(i.full_name for i in obj.DFO_resource_manager.all()),
            obj.funding_recipient,
            obj.first_nation.name if obj.first_nation else None,
            obj.contact.full_name if obj.contact else None,
            obj.contact_role,
            ", ".join(i.full_name for i in obj.DFO_technicians.all()),
            obj.contractor,
            obj.contractor_contact,
            ", ".join(i.name for i in obj.partner.all()),
            ", ".join(i.full_name for i in obj.partner_contact.all()),

            obj.agreement_database,
            obj.agreement_comment,
            ", ".join(i.name for i in obj.funding_sources.all()),
            obj.other_funding_sources,
            obj.agreement_type,
            obj.organization_program.name if obj.organization_program else None,

            obj.date_last_modified,
            obj.last_modified_by.get_full_name() if obj.last_modified_by else None,

        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

        ##########
        # RIVERS #
        for river_tmp in obj.river.all():
            river_num += 1
            river_rows = [
                obj.agreement_number,
                obj.project_number,
                obj.name,
                ", ".join(i.full_name for i in obj.DFO_project_authority.all()),
                obj.biological_process_type_1,
                obj.biological_process_type_2,
                ", ".join(i.name for i in obj.biological_process_type_3.all()),
                ", ".join(i.name for i in obj.activity_type_1.all()),
                ", ".join(i.name for i in obj.activity_type_2.all()),
                ", ".join(i.name for i in obj.activity_type_3.all()),
                obj.area,
                obj.first_nation.name if obj.first_nation else None,
                obj.funding_recipient,
                obj.organization_program.name if obj.organization_program else None,
                ", ".join(i.name for i in obj.hatchery_name.all()),
                river_tmp.name,
                river_tmp.latitude,
                river_tmp.longitude,
                river_tmp.sub_district_area.name if river_tmp.sub_district_area else None,
                river_tmp.species if river_tmp.species else None,
                river_tmp.stock_management_unit if river_tmp.stock_management_unit else None,
                river_tmp.cu_index.name if river_tmp.cu_index else None,
                river_tmp.cu_name.name if river_tmp.cu_name else None,
                river_tmp.popid if river_tmp.popid else None,
                river_tmp.du.name if river_tmp.du else None,
                river_tmp.du_number if river_tmp.du_number else None,
                river_tmp.watershed.name if river_tmp.watershed else None,

            ]
            for col_num, cell_value in enumerate(river_rows, 1):
                cell = river_worksheet.cell(row=river_num, column=col_num)
                cell.value = cell_value

        ##################
        # FUNDING YEARS #
        for fund in funding_year:
            fund_num += 1
            fund_rows = [
                obj.agreement_number,
                obj.project_number,
                obj.name,
                fund.funding_year,
                ", ".join(i.name for i in obj.funding_sources.all()),
                ", ".join(i.full_name for i in obj.DFO_project_authority.all()),
                obj.first_nation.name if obj.first_nation else None,
                ", ".join(i.name for i in obj.river.all()),
                ", ".join(i.species for i in obj.river.all()),
                all_sub_tmp,
                ", ".join(i.stock_management_unit if i.stock_management_unit else "" for i in obj.river.all()),
                all_cui_tmp,
                all_cu_tmp,
                all_du_tmp,
                ", ".join(i.du_number if i.du_number else "" for i in obj.river.all()),
                all_wtr_tmp,
                obj.funding_recipient,
                obj.biological_process_type_1,
                obj.biological_process_type_2,
                ", ".join(i.name for i in obj.biological_process_type_3.all()),
                ", ".join(i.name for i in obj.activity_type_1.all()),
                ", ".join(i.name for i in obj.activity_type_2.all()),
                ", ".join(i.name for i in obj.activity_type_3.all()),
                obj.area,
                fund.agreement_cost,
                fund.project_cost,
                fund.date_last_modified,
                fund.last_modified_by.get_full_name() if fund.last_modified_by else None,
            ]
            for col_num, cell_value in enumerate(fund_rows, 1):
                cell = funding_worksheet.cell(row=fund_num, column=col_num)
                cell.value = cell_value

        ##########################
        # PROJECT CERTIFIED ROWS #
        project_certified = models.ProjectCertified.objects.filter(project=obj.id)
        for certified in project_certified:
            cert_num += 1
            cert_rows = [
                obj.agreement_number,
                obj.project_number,
                obj.name,
                ", ".join(i.full_name for i in obj.DFO_project_authority.all()),
                obj.biological_process_type_1,
                obj.biological_process_type_2,
                ", ".join(i.name for i in obj.biological_process_type_3.all()),
                ", ".join(i.name for i in obj.activity_type_1.all()),
                ", ".join(i.name for i in obj.activity_type_2.all()),
                ", ".join(i.name for i in obj.activity_type_3.all()),
                obj.area,
                ", ".join(i.name for i in obj.river.all()),
                ", ".join(i.species for i in obj.river.all()),
                all_sub_tmp,
                ", ".join(i.stock_management_unit if i.stock_management_unit else "" for i in obj.river.all()),
                all_cui_tmp,
                all_du_tmp,
                ", ".join(i.du_number if i.du_number else "" for i in obj.river.all()),
                all_cu_tmp,
                all_wtr_tmp,
                certified.certified_date,
                certified.certified_by.get_full_name() if certified.certified_by else None,
            ]
            for col_num, cell_value in enumerate(cert_rows, 1):
                cell = certified_worksheet.cell(row=cert_num, column=col_num)
                cell.value = cell_value

        ##################
        # ACTIVITY OUTCOME ROWS #
        activity_filter = models.ActivitiesAndOutcomes.objects.filter(project=obj.id)
        for activity in activity_filter:
            obj_num += 1
            if activity.river and activity.river.cu_name:
                obj_cu_tmp = activity.river.cu_name.name
            else:
                obj_cu_tmp = None
            if activity.river and activity.river.watershed:
                obj_wtr_tmp = activity.river.watershed.name
            else:
                obj_wtr_tmp = None
            if activity.river and activity.river.sub_district_area:
                obj_sub_tmp = activity.river.sub_district_area.name
            else:
                obj_sub_tmp = None
            if activity.river and activity.river.cu_index:
                obj_cui_tmp = activity.river.cu_index.name
            else:
                obj_cui_tmp = None
            if activity.river and activity.river.du:
                obj_du_tmp = activity.river.du.name
            else:
                obj_du_tmp = None
            activity_outcome_row = [
                obj.agreement_number,
                activity.unique_activity_outcome_number,
                obj.project_number,
                obj.name,
                ", ".join(i.funding_year for i in funding_year.all()),
                ", ".join(i.name for i in obj.funding_sources.all()),
                ", ".join(i.full_name for i in obj.DFO_project_authority.all()),
                obj.biological_process_type_1 if obj else None,
                obj.biological_process_type_2 if obj else None,
                ", ".join(i.name for i in obj.biological_process_type_3.all()),
                ", ".join(i.name for i in obj.activity_type_1.all()),
                ", ".join(i.name for i in obj.activity_type_2.all()),
                ", ".join(i.name for i in obj.activity_type_3.all()),
                obj.area if obj else None,
                activity.river.name if activity.river else None,
                activity.river.species if activity.river else None,
                obj_sub_tmp,
                activity.river.stock_management_unit if activity.river else None,
                obj_cui_tmp,
                obj_du_tmp,
                activity.river.du_number if activity.river else None,
                obj_cu_tmp,
                obj_wtr_tmp,
                obj.first_nation.name if obj.first_nation else None,
                obj.funding_recipient,
                obj.organization_program.name if obj.organization_program else None,
                ", ".join(i.name for i in obj.hatchery_name.all()),
                activity.task_description,
                activity.element_title,
                activity.activity_title,
                activity.pst_requirement,
                ", ".join(i.name for i in activity.activity_outcome_category.all()),
                activity.sil_requirement,
                activity.expected_results,
                activity.dfo_report,
                activity.outcomes_comment,
                activity.date_last_modified,
                activity.last_modified_by.get_full_name() if activity.last_modified_by else None,
            ]

            for col_num, cell_value in enumerate(activity_outcome_row, 1):
                cell = activity_outcome_worksheet.cell(row=obj_num, column=col_num)
                cell.value = cell_value

            #######################
            # SAMPLE OUTCOME ROWS #
            sample_outcomes = models.SampleOutcome.objects.filter(activity_outcome=activity.id)
            for sample in sample_outcomes:
                sampo_num += 1
                sample_rows = [
                    obj.agreement_number,
                    obj.project_number,
                    obj.name,
                    ", ".join(i.funding_year for i in funding_year.all()),
                    ", ".join(i.name for i in obj.funding_sources.all()),
                    ", ".join(i.full_name for i in obj.DFO_project_authority.all()),
                    obj.biological_process_type_1 if obj else None,
                    obj.biological_process_type_2 if obj else None,
                    ", ".join(i.name for i in obj.biological_process_type_3.all()),
                    ", ".join(i.name for i in obj.activity_type_1.all()),
                    ", ".join(i.name for i in obj.activity_type_2.all()),
                    ", ".join(i.name for i in obj.activity_type_3.all()),
                    obj.area if obj else None,
                    activity.river.name if activity.river else None,
                    activity.river.species if activity.river else None,
                    obj_sub_tmp,
                    activity.river.stock_management_unit if activity.river else None,
                    obj_cui_tmp,
                    obj_du_tmp,
                    activity.river.du_number if activity.river else None,
                    obj_cu_tmp,
                    obj_wtr_tmp,
                    obj.first_nation.name if obj.first_nation else None,
                    obj.funding_recipient,
                    obj.organization_program.name if obj.organization_program else None,
                    ", ".join(i.name for i in obj.hatchery_name.all()),
                    sample.unique_activity_outcome_number,
                    sample.sampling_outcome,
                    sample.outcome_delivered,
                    sample.outcome_quality,
                    sample.sample_outcome_comment,
                    sample.date_last_modified,
                    sample.last_modified_by.get_full_name() if sample.last_modified_by else None,
                ]
                for col_num, cell_value in enumerate(sample_rows, 1):
                    cell = sample_outcome_worksheet.cell(row=sampo_num, column=col_num)
                    cell.value = cell_value
            #######################
            # REPORT OUTCOME ROWS #
            report_outcomes = models.ReportOutcome.objects.filter(activity_outcome=activity.id)
            for report in report_outcomes:
                repo_num += 1
                report_rows = [
                    report.activity_outcome.project.agreement_number if report.activity_outcome.project else None,
                    obj.project_number,
                    obj.name,
                    ", ".join(i.funding_year for i in funding_year.all()),
                    ", ".join(i.name for i in obj.funding_sources.all()),
                    ", ".join(i.full_name for i in obj.DFO_project_authority.all()),
                    obj.biological_process_type_1 if obj else None,
                    obj.biological_process_type_2 if obj else None,
                    ", ".join(i.name for i in obj.biological_process_type_3.all()),
                    ", ".join(i.name for i in obj.activity_type_1.all()),
                    ", ".join(i.name for i in obj.activity_type_2.all()),
                    ", ".join(i.name for i in obj.activity_type_3.all()),
                    obj.area if obj else None,
                    activity.river.name if activity.river else None,
                    activity.river.species if activity.river else None,
                    obj_sub_tmp,
                    activity.river.stock_management_unit if activity.river else None,
                    obj_cui_tmp,
                    obj_du_tmp,
                    activity.river.du_number if activity.river else None,
                    obj_cu_tmp,
                    obj_wtr_tmp,
                    obj.first_nation.name if obj.first_nation else None,
                    obj.funding_recipient,
                    obj.organization_program.name if obj.organization_program else None,
                    ", ".join(i.name for i in obj.hatchery_name.all()),
                    report.unique_activity_outcome_number,
                    report.reporting_outcome,
                    report.task_id,
                    report.task_name,
                    report.task_description,
                    report.task_met,
                    report.site,
                    report.report_link.document_name if report.report_link else None,
                    report.report_outcome_comment,
                    report.reporting_outcome_metric,
                    report.reporting_outcome_metric_unit,
                    report.data_sharing,
                    report.date_last_modified,
                    report.last_modified_by.get_full_name() if report.last_modified_by else None,
                ]
                for col_num, cell_value in enumerate(report_rows, 1):
                    cell = reporting_outcome_worksheet.cell(row=repo_num, column=col_num)
                    cell.value = cell_value

        #############
        # DATA ROWS #
        data_filter = models.Data.objects.filter(project=obj.id)
        for data in data_filter:
            if data.river and data.river.cu_name:
                data_cu_tmp = data.river.cu_name.name
            else:
                data_cu_tmp = None
            if data.river and data.river.watershed:
                data_wtr_tmp = data.river.watershed.name
            else:
                data_wtr_tmp = None
            if data.river and data.river.sub_district_area:
                data_sub_tmp = data.river.sub_district_area.name
            else:
                data_sub_tmp = None
            if data.river and data.river.cu_index:
                data_cui_tmp = data.river.cu_index.name
            else:
                data_cui_tmp = None
            if data.river and data.river.du:
                data_du_tmp = data.river.du.name
            else:
                data_du_tmp = None
            data_num += 1
            data_row = [
                data.project.agreement_number if data.project else None,
                obj.project_number,
                obj.name,
                ", ".join(i.funding_year for i in funding_year.all()),
                ", ".join(i.name for i in obj.funding_sources.all()),
                ", ".join(i.full_name for i in obj.DFO_project_authority.all()),
                obj.biological_process_type_1 if obj else None,
                obj.biological_process_type_2 if obj else None,
                ", ".join(i.name for i in obj.biological_process_type_3.all()),
                ", ".join(i.name for i in obj.activity_type_1.all()),
                ", ".join(i.name for i in obj.activity_type_2.all()),
                ", ".join(i.name for i in obj.activity_type_3.all()),
                obj.area if obj else None,
                data.river.name if data.river else None,
                data.river.species if data.river else None,
                data_sub_tmp,
                data.river.stock_management_unit if data.river else None,
                data_cui_tmp,
                data_du_tmp,
                data.river.du_number if data.river else None,
                data_cu_tmp,
                data_wtr_tmp,
                obj.first_nation.name if obj.first_nation else None,
                obj.funding_recipient,
                obj.organization_program.name if obj.organization_program else None,
                ", ".join(i.name for i in obj.hatchery_name.all()),
                data.samples_collected,
                data.samples_collected_comment,
                data.samples_collected_database,
                data.shared_drive,
                data.sample_entered_database,
                data.data_quality_check,
                data.sample_format,
                data.data_products,
                data.data_products_database,
                data.data_products_comment,
                data.date_last_modified,
                data.last_modified_by.get_full_name() if data.last_modified_by else None,
            ]
            for col_num, cell_value in enumerate(data_row, 1):
                cell = data_worksheet.cell(row=data_num, column=col_num)
                cell.value = cell_value

        ###############
        # METHOD ROWS #
        method_filter = models.Method.objects.filter(project=obj.id)
        for method in method_filter:
            meth_num += 1
            method_row = [
                method.project.agreement_number if method.project else None,
                method.unique_method_number,
                obj.project_number,
                obj.name,
                ", ".join(i.funding_year for i in funding_year.all()),
                ", ".join(i.name for i in obj.funding_sources.all()),
                ", ".join(i.full_name for i in obj.DFO_project_authority.all()),
                obj.biological_process_type_1,
                obj.biological_process_type_2,
                ", ".join(i.name for i in obj.biological_process_type_3.all()),
                ", ".join(i.name for i in obj.activity_type_1.all()),
                ", ".join(i.name for i in obj.activity_type_2.all()),
                ", ".join(i.name for i in obj.activity_type_3.all()),
                obj.area,
                ", ".join(i.name for i in obj.river.all()),
                ", ".join(i.species for i in obj.river.all()),
                all_sub_tmp,
                ", ".join(i.stock_management_unit if i.stock_management_unit else "" for i in obj.river.all()),
                all_cui_tmp,
                all_du_tmp,
                ", ".join(i.du_number if i.du_number else "" for i in obj.river.all()),
                all_cu_tmp,
                all_wtr_tmp,
                obj.first_nation.name if obj.first_nation else None,
                obj.funding_recipient,
                obj.organization_program.name if obj.organization_program else None,
                ", ".join(i.name for i in obj.hatchery_name.all()),
                method.field_work_method_type,
                method.planning_method_type,
                method.sample_processing_method_type,
                method.scale_processing_location,
                method.otolith_processing_location,
                method.DNA_processing_location,
                method.heads_processing_location,
                method.instrument_data_processing_location,
                method.date_last_modified,
                method.last_modified_by.get_full_name() if method.last_modified_by else None,
            ]
            for col_num, cell_value in enumerate(method_row, 1):
                cell = method_worksheet.cell(row=meth_num, column=col_num)
                cell.value = cell_value

        ########################
        # METHOD DOCUMENT ROWS #
            method_documents = models.MethodDocument.objects.filter(method=method.id)
            for method_doc in method_documents:
                methd_num += 1
                method_doc_rows = [
                    obj.agreement_number,
                    method_doc.unique_method_number,
                    obj.project_number,
                    obj.name,
                    ", ".join(i.funding_year for i in funding_year.all()),
                    ", ".join(i.name for i in obj.funding_sources.all()),
                    ", ".join(i.full_name for i in obj.DFO_project_authority.all()),
                    obj.biological_process_type_1,
                    obj.biological_process_type_2,
                    ", ".join(i.name for i in obj.biological_process_type_3.all()),
                    ", ".join(i.name for i in obj.activity_type_1.all()),
                    ", ".join(i.name for i in obj.activity_type_2.all()),
                    ", ".join(i.name for i in obj.activity_type_3.all()),
                    obj.area,
                    ", ".join(i.name for i in obj.river.all()),
                    ", ".join(i.species for i in obj.river.all()),
                    all_sub_tmp,
                    ", ".join(i.stock_management_unit if i.stock_management_unit else "" for i in obj.river.all()),
                    all_cui_tmp,
                    all_du_tmp,
                    ", ".join(i.du_number if i.du_number else "" for i in obj.river.all()),
                    all_cu_tmp,
                    all_wtr_tmp,
                    obj.first_nation.name if obj.first_nation else None,
                    obj.funding_recipient,
                    obj.organization_program.name if obj.organization_program else None,
                    ", ".join(i.name for i in obj.hatchery_name.all()),
                    method_doc.method_document_type,
                    method_doc.authors,
                    method_doc.publication_year,
                    method_doc.title,
                    method_doc.reference_number,
                    method_doc.document_link,
                    method_doc.date_last_modified,
                    method_doc.last_modified_by.get_full_name() if method_doc.last_modified_by else None,
                ]
                for col_num, cell_value in enumerate(method_doc_rows, 1):
                    cell = method_document_worksheet.cell(row=methd_num, column=col_num)
                    cell.value = cell_value

    ###############
    # REPORTS ROWS #
        reports_filter = models.Reports.objects.filter(project=obj.id)
        for report in reports_filter:
            rep_num += 1
            report_row = [
                report.project.agreement_number if report.project else None,
                obj.project_number,
                obj.name,
                ", ".join(i.funding_year for i in funding_year.all()),
                ", ".join(i.name for i in obj.funding_sources.all()),
                ", ".join(i.full_name for i in obj.DFO_project_authority.all()),
                obj.biological_process_type_1,
                obj.biological_process_type_2,
                ", ".join(i.name for i in obj.biological_process_type_3.all()),
                ", ".join(i.name for i in obj.activity_type_1.all()),
                ", ".join(i.name for i in obj.activity_type_2.all()),
                ", ".join(i.name for i in obj.activity_type_3.all()),
                obj.area,
                ", ".join(i.name for i in obj.river.all()),
                ", ".join(i.species for i in obj.river.all()),
                all_sub_tmp,
                ", ".join(i.stock_management_unit if i.stock_management_unit else ""for i in obj.river.all()),
                all_cui_tmp,
                all_du_tmp,
                ", ".join(i.du_number if i.du_number else "" for i in obj.river.all()),
                all_cu_tmp,
                all_wtr_tmp,
                obj.first_nation.name if obj.first_nation else None,
                obj.funding_recipient,
                obj.organization_program.name if obj.organization_program else None,
                ", ".join(i.name for i in obj.hatchery_name.all()),
                report.report_timeline,
                report.report_type,
                report.document_name,
                report.document_author,
                report.document_reference_information,
                report.document_link,
                report.published,
                report.date_last_modified,
                report.last_modified_by.get_full_name() if report.last_modified_by else None,
            ]
            for col_num, cell_value in enumerate(report_row, 1):
                cell = report_worksheet.cell(row=rep_num, column=col_num)
                cell.value = cell_value

    workbook.save(response)
    return response


def export_project(request):
    project = models.Project.objects.all()
    project_filter = filters.ProjectFilter(request.GET, queryset=project).qs
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-Projects.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Projects'

    columns = [
        'Agreement Number',
        'Project Number',

        'Agreement History',
        'Project Name',
        'Project Description',
        'Funding Year',
        'Start Date',
        'End Date',

        'Area',
        'Other Species',
        'Ecosystem Type',
        'Lake System(s)',
        'Other Site Info',

        'Salmon Life Stage',
        'Aquaculture License Number',
        'Water License Number',
        'Hatchery Name',
        'DFO Tenure',

        'Project Stage',
        'Biological Process Type - Level 1',
        'Activity Type - level 1',
        'Monitoring Approach',
        'Biological Process Type - level 3',
        'Biological Process Type - level 2',
        'Activity Type - level 2',
        'Activity Type - level 2',
        'Project Purpose',
        'Category Comments',

        'Link to other DFO Programs',
        'Linked DFO Program Project Reference',
        'Link to other Government Departments',
        'Policy Connections',

        'DFO Project Authority',
        'DFO Area Chief',
        'DFO IAA',
        'DFO Resource Manager',
        'Funding Recipient',
        'First Nation/Tribal Council',
        'Contact',
        'Contact Role',
        'DFO Technicians/Biologists',
        'Contractors',
        'Contractor Contact',
        'Partner',
        'Partner Contact',

        'Agreement Database',
        'Agreement Comment',
        'Funding Sources',
        'Other Funding Sources',
        'Agreement Type',
        'Primary Organization or Program',

        'date_last_modified',
        'last_modified_by',


    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = 15
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.value = column_title

    ##############################
    # Rivers SHEET/HEADER #
    river_worksheet = workbook.create_sheet("Rivers", 2)
    river_columns = [
        'Agreement Number',
        'Project Number',
        'Project Name',
        'Project Authority',
        'Biological Process Type - Level 1',
        'Biological Process Type - level 2',
        'Biological Process Type - level 3',
        'Activity Type - level 1',
        'Activity Type - level 2',
        'Activity Type - level 3',
        'Area',
        'First Nations',
        'Funding Recipient',
        'Primary Organization or Program',
        'Hatchery',
        'Name',
        'Latitude',
        'Longitude',
        'Sub District Area',
        'Species',
        'Stock Management Unit',
        'CU index',
        'CU Name',
        'Pop ID',
        'DU',
        'DU Number',
        'Watershed',
    ]
    row_num = 1
    for col_num, column_title in enumerate(river_columns, 1):
        river_worksheet.column_dimensions[get_column_letter(col_num)].width = 15
        cell = river_worksheet.cell(row=row_num, column=col_num)
        cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.value = column_title


    ###########
    # FUNDING #
    funding_worksheet = workbook.create_sheet("Funding", 1)
    funding_columns = [
        'Agreement Number',
        'Project Number',
        'Project Name',
        'Funding Year',
        'Funding Source',
        'Project Authority',
        'First Nation',
        'Location / River',
        'Species',
        'Sub District Area',
        'Stock Management Unit',
        'CU Index',
        'CU Name',
        'DU',
        'DU Number',
        'Watershed',
        'Funding Recipient',
        'Biological Process Type - Level 1',
        'Biological Process Type - Level 2',
        'Biological Process Type - Level 3',
        'Activity Type - level 1',
        'Activity Type - level 2',
        'Activity Type - level 3',
        'Area',
        'Agreement Cost',
        'Project Cost',
        'date_last_modified',
        'last_modified_by',
    ]
    row_num = 1
    for col_num, column_title in enumerate(funding_columns, 1):
        funding_worksheet.column_dimensions[get_column_letter(col_num)].width = 15
        cell = funding_worksheet.cell(row=row_num, column=col_num)
        cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.value = column_title

    #############
    # CERTIFIED #
    certified_worksheet = workbook.create_sheet("Project Certified", 3)
    certified_columns = [
        'Agreement Number',
        'Project Number',
        'Project Name',
        'Biological Process Type - Level 1',
        'Biological Process Type - level 2',
        'Biological Process Type - level 3',
        'Activity Type - level 1',
        'Activity Type - level 2',
        'Activity Type - level 3',
        'Area',
        'Location/River',
        'Species',
        'Sub District Area',
        'Stock Management Unit',
        'CU index',
        'DU',
        'DU Number',
        'CU Name',
        'Watershed',
        'Certified Date',
        'Certified By',
    ]
    row_num = 1
    for col_num, column_title in enumerate(certified_columns, 1):
        certified_worksheet.column_dimensions[get_column_letter(col_num)].width = 15
        cell = certified_worksheet.cell(row=row_num, column=col_num)
        cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.value = column_title

    fund_num = 1
    cert_num = 1
    river_num = 1

    # PROJECT ROWS #
    for obj in project_filter:
        all_cu_tmp = ""
        all_wtr_tmp = ""
        all_sub_tmp = ""
        all_cui_tmp = ""
        all_du_tmp = ""
        for i in obj.river.all():
            if i.cu_name:
                all_cu_tmp += str(i.cu_name.name + ", ")
            if i.watershed:
                all_wtr_tmp += str(i.watershed.name + ", ")
            if i.sub_district_area:
                all_sub_tmp += str(i.sub_district_area.name + ", ")
            if i.cu_index:
                all_cui_tmp += str(i.cu_index.name + ", ")
            if i.du:
                all_du_tmp += str(i.du.name + ", ")
        row_num += 1
        funding_year = models.FundingYears.objects.filter(project=obj.id)
        row = [
            obj.agreement_number,
            obj.project_number,

            ", ".join(i.name for i in obj.agreement_history.all()),
            obj.name,
            obj.project_description,
            ", ".join(i.funding_year for i in funding_year),
            obj.start_date,
            obj.end_date,

            obj.area,
            obj.other_species,
            ", ".join(i.name for i in obj.ecosystem_type.all()),
            ", ".join(i.name for i in obj.lake_system.all()),
            obj.other_site_info,

            ", ".join(i.name for i in obj.salmon_life_stage.all()),
            obj.aquaculture_license_number,
            obj.water_license_number,
            ", ".join(i.name for i in obj.hatchery_name.all()),
            obj.DFO_tenure,

            obj.project_stage,
            obj.biological_process_type_1,
            ", ".join(i.name for i in obj.activity_type_1.all()),
            ", ".join(i.name for i in obj.monitoring_approach.all()),
            ", ".join(i.name for i in obj.biological_process_type_3.all()),
            obj.biological_process_type_2,
            ", ".join(i.name for i in obj.activity_type_2.all()),
            ", ".join(i.name for i in obj.activity_type_3.all()),
            ", ".join(i.name for i in obj.project_purpose.all()),
            obj.category_comments,

            obj.DFO_link,
            obj.DFO_program_reference,
            obj.government_organization,
            obj.policy_connection,

            ", ".join(i.full_name for i in obj.DFO_project_authority.all()),
            ", ".join(i.full_name for i in obj.DFO_area_chief.all()),
            ", ".join(i.full_name for i in obj.DFO_IAA.all()),
            ", ".join(i.full_name for i in obj.DFO_resource_manager.all()),
            obj.funding_recipient,
            obj.first_nation.name if obj.first_nation else None,
            obj.contact.full_name if obj.contact else None,
            obj.contact_role,
            ", ".join(i.full_name for i in obj.DFO_technicians.all()),
            obj.contractor,
            obj.contractor_contact,
            ", ".join(i.name for i in obj.partner.all()),
            ", ".join(i.full_name for i in obj.partner_contact.all()),

            obj.agreement_database,
            obj.agreement_comment,
            ", ".join(i.name for i in obj.funding_sources.all()),
            obj.other_funding_sources,
            obj.agreement_type,
            obj.organization_program.name if obj.organization_program else None,

            obj.date_last_modified,
            obj.last_modified_by.get_full_name() if obj.last_modified_by else None,

        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

        ##############
        # RIVER ROWS #
        for river_tmp in obj.river.all():
            river_num += 1
            river_rows = [
                obj.agreement_number,
                obj.project_number,
                obj.name,
                ", ".join(i.full_name for i in obj.DFO_project_authority.all()),
                obj.biological_process_type_1,
                obj.biological_process_type_2,
                ", ".join(i.name for i in obj.biological_process_type_3.all()),
                ", ".join(i.name for i in obj.activity_type_1.all()),
                ", ".join(i.name for i in obj.activity_type_2.all()),
                ", ".join(i.name for i in obj.activity_type_3.all()),
                obj.area,
                obj.first_nation.name if obj.first_nation else None,
                obj.funding_recipient,
                obj.organization_program.name if obj.organization_program else None,
                ", ".join(i.name for i in obj.hatchery_name.all()),
                river_tmp.name if river_tmp.name else None,
                river_tmp.latitude if river_tmp.latitude else None,
                river_tmp.longitude if river_tmp.longitude else None,
                river_tmp.sub_district_area.name if river_tmp.sub_district_area else None,
                river_tmp.species if river_tmp.species else None,
                river_tmp.stock_management_unit if river_tmp.stock_management_unit else None,
                river_tmp.cu_index.name if river_tmp.cu_index else None,
                river_tmp.cu_name.name if river_tmp.cu_name else None,
                river_tmp.popid if river_tmp.popid else None,
                river_tmp.du.name if river_tmp.du else None,
                river_tmp.du_number if river_tmp.du_number else None,
                river_tmp.watershed.name if river_tmp.watershed else None,
            ]
            for col_num, cell_value in enumerate(river_rows, 1):
                cell = river_worksheet.cell(row=river_num, column=col_num)
                cell.value = cell_value

            # FUNDING ROWS #
        funding_years = models.FundingYears.objects.filter(project=obj.id)
        for fund in funding_years:
            fund_num += 1
            fund_rows = [
                obj.agreement_number,
                obj.project_number,
                obj.name,
                fund.funding_year,
                ", ".join(i.name for i in obj.funding_sources.all()),
                ", ".join(i.full_name for i in obj.DFO_project_authority.all()),
                obj.first_nation.name if obj.first_nation else None,
                ", ".join(i.name for i in obj.river.all()),
                ", ".join(i.species for i in obj.river.all()),
                all_sub_tmp,
                ", ".join(i.stock_management_unit if i.stock_management_unit else "" for i in obj.river.all()),
                all_cui_tmp,
                all_cu_tmp,
                all_du_tmp,
                ", ".join(i.du_number if i.du_number else "" for i in obj.river.all()),
                all_wtr_tmp,
                obj.funding_recipient,
                obj.biological_process_type_1,
                obj.biological_process_type_2,
                ", ".join(i.name for i in obj.biological_process_type_3.all()),
                ", ".join(i.name for i in obj.activity_type_1.all()),
                ", ".join(i.name for i in obj.activity_type_2.all()),
                ", ".join(i.name for i in obj.activity_type_3.all()),
                obj.area,
                fund.agreement_cost,
                fund.project_cost,
                fund.date_last_modified,
                fund.last_modified_by.get_full_name() if fund.last_modified_by else None,
            ]
            for col_num, cell_value in enumerate(fund_rows, 1):
                cell = funding_worksheet.cell(row=fund_num, column=col_num)
                cell.value = cell_value

        # CERTIFIED ROWS #
        project_certified = models.ProjectCertified.objects.filter(project=obj.id)
        for certified in project_certified:
            cert_num += 1
            cert_rows = [
                obj.agreement_number,
                obj.project_number,
                obj.name,
                ", ".join(i.full_name for i in obj.DFO_project_authority.all()),
                obj.biological_process_type_1,
                obj.biological_process_type_2,
                ", ".join(i.name for i in obj.biological_process_type_3.all()),
                ", ".join(i.name for i in obj.activity_type_1.all()),
                ", ".join(i.name for i in obj.activity_type_2.all()),
                ", ".join(i.name for i in obj.activity_type_3.all()),
                obj.area,
                ", ".join(i.name for i in obj.river.all()),
                ", ".join(i.species for i in obj.river.all()),
                all_sub_tmp,
                ", ".join(i.stock_management_unit if i.stock_management_unit else "" for i in obj.river.all()),
                all_cui_tmp,
                all_du_tmp,
                ", ".join(i.du_number if i.du_number else "" for i in obj.river.all()),
                all_cu_tmp,
                all_wtr_tmp,
                certified.certified_date,
                certified.certified_by.get_full_name() if certified.certified_by else None,
            ]
            for col_num, cell_value in enumerate(cert_rows, 1):
                cell = certified_worksheet.cell(row=cert_num, column=col_num)
                cell.value = cell_value

    workbook.save(response)

    return response


def export_activity_outcome(request):
    activity = models.ActivitiesAndOutcomes.objects.all()
    activity_filter = filters.ActivitiesAndOutcomesFilter(request.GET, queryset=activity).qs
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-Activities&Outcomes.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Activities & Outcomes'

    columns = [
        'Agreement Number',
        'Unique Activity Outcome Number',
        'Project Number',
        'Project Name',
        'Funding Year',
        'Funding Source',
        'Project Authority',
        'Biological Process Type - Level 1',
        'Biological Process Type - level 2',
        'Biological Process Type - level 3',
        'Activity Type - level 1',
        'Activity Type - level 2',
        'Activity Type - level 3',
        'Area',
        'Location/River',
        'Species',
        'Sub District Area',
        'Stock Management Unit',
        'CU index',
        'DU',
        'DU Number',
        'CU Name',
        'Watershed',
        'First Nations',
        'Funding Recipient',
        'Primary Organization or Program',
        'Hatchery',
        'Task Description',
        'Element Title',
        'Activity Title',
        'PST Requirement Identified?',
        'Activity & Outcome Category',
        'SIL Requirement',
        'Expected Result(s)',
        'Products/Reports to Provide DFO',
        'Outcome Comment',
        'date_last_modified',
        'last_modified_by',

    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = 15
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.value = column_title

    # Sample Outcome #
    sample_outcome_worksheet = workbook.create_sheet("Sample Outcomes", 1)
    sample_outcomes_columns = [
        'Agreement Number',
        'Project Number',
        'Project Name',
        'Funding Year',
        'Funding Source',
        'Project Authority',
        'Biological Process Type - Level 1',
        'Biological Process Type - level 2',
        'Biological Process Type - level 3',
        'Activity Type - level 1',
        'Activity Type - level 2',
        'Activity Type - level 3',
        'Area',
        'Location/River',
        'Species',
        'Sub District Area',
        'Stock Management Unit',
        'CU index',
        'DU',
        'DU Number',
        'CU Name',
        'Watershed',
        'First Nations',
        'Funding Recipient',
        'Primary Organization or Program',
        'Hatchery',
        'Unique Activity and Outcome Number',
        'Sampling Outcome',
        'Were outcomes delivered?',
        'Quality of Outcome',
        'Sample Outcome Comment',
        'date_last_modified',
        'last_modified_by',
    ]
    row_num = 1

    for col_num, column_title in enumerate(sample_outcomes_columns, 1):
        sample_outcome_worksheet.column_dimensions[get_column_letter(col_num)].width = 15
        cell = sample_outcome_worksheet.cell(row=row_num, column=col_num)
        cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.value = column_title

    # Reporting Outcome #
    reporting_outcome_worksheet = workbook.create_sheet("Report Outcomes", 2)
    report_outcomes_columns = [
        'Agreement Number',
        'Project Number',
        'Project Name',
        'Funding Year',
        'Funding Source',
        'Project Authority',
        'Biological Process Type - Level 1',
        'Biological Process Type - level 2',
        'Biological Process Type - level 3',
        'Activity Type - level 1',
        'Activity Type - level 2',
        'Activity Type - level 3',
        'Area',
        'Location/River',
        'Species',
        'Sub District Area',
        'Stock Management Unit',
        'CU index',
        'DU',
        'DU Number',
        'CU Name',
        'Watershed',
        'First Nations',
        'Funding Recipient',
        'Primary Organization or Program',
        'Hatchery',
        'Unique Activity and Outcome Number',
        'Reporting Outcome',
        'Task ID',
        'Task Name',
        'Task Description',
        'Task Met',
        'Site',
        'Report Link',
        'Reporting Outcome Comment',
        'Reporting Outcome Metric',
        'Reporting Outcome Metric Unit',
        'Data Sharing',
        'date_last_modified',
        'last_modified_by',
    ]
    row_num = 1

    for col_num, column_title in enumerate(report_outcomes_columns, 1):
        reporting_outcome_worksheet.column_dimensions[get_column_letter(col_num)].width = 15
        cell = reporting_outcome_worksheet.cell(row=row_num, column=col_num)
        cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.value = column_title

    samp_num = 1
    row_num = 1
    rep_num = 1
    for obj in activity_filter:
        row_num += 1
        proj = models.Project.objects.get(pk=obj.project.id)
        if obj.river and obj.river.cu_name:
            cu_tmp = obj.river.cu_name.name
        else:
            cu_tmp = None
        if obj.river and obj.river.watershed:
            wtr_tmp = obj.river.watershed.name
        else:
            wtr_tmp = None
        if obj.river and obj.river.sub_district_area:
            sub_tmp = obj.river.sub_district_area.name
        else:
            sub_tmp = None
        if obj.river and obj.river.du:
            du_tmp = obj.river.du.name
        else:
            du_tmp = None
        if obj.river and obj.river.cu_index:
            cui_tmp = obj.river.cu_index.name
        else:
            cui_tmp = None
        funding_year = models.FundingYears.objects.filter(project=obj.project.id)
        row = [
            proj.agreement_number if proj else None,
            obj.unique_activity_outcome_number,
            proj.project_number if proj else None,
            proj.name if proj else None,
            ", ".join(i.funding_year for i in funding_year.all()),
            ", ".join(i.name for i in proj.funding_sources.all()),
            ", ".join(i.full_name for i in proj.DFO_project_authority.all()),
            proj.biological_process_type_1 if proj else None,
            proj.biological_process_type_2 if proj else None,
            ", ".join(i.name for i in proj.biological_process_type_3.all()),
            ", ".join(i.name for i in proj.activity_type_1.all()),
            ", ".join(i.name for i in proj.activity_type_2.all()),
            ", ".join(i.name for i in proj.activity_type_3.all()),
            proj.area if proj else None,
            obj.river.name if obj.river else None,
            obj.river.species if obj.river else None,
            sub_tmp,
            obj.river.stock_management_unit if obj.river else None,
            cui_tmp,
            du_tmp,
            obj.river.du_number if obj.river else None,
            cu_tmp,
            wtr_tmp,
            proj.first_nation.name if proj.first_nation else None,
            proj.funding_recipient if proj else None,
            proj.organization_program.name if proj.organization_program else None,
            ", ".join(i.name for i in proj.hatchery_name.all()),
            obj.task_description,
            obj.element_title,
            obj.activity_title,
            obj.pst_requirement,
            ", ".join(i.name for i in obj.activity_outcome_category.all()),
            obj.sil_requirement,
            obj.expected_results,
            obj.dfo_report,
            obj.outcomes_comment,
            obj.date_last_modified,
            obj.last_modified_by.get_full_name() if obj.last_modified_by else None,

        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

        ##################
        # SAMPLE OUTCOME #
        sample_outcomes = models.SampleOutcome.objects.filter(activity_outcome=obj.id)
        for sample in sample_outcomes:
            samp_num += 1
            sample_rows = [
                proj.agreement_number if proj else None,
                proj.project_number if proj else None,
                proj.name if proj else None,
                ", ".join(i.funding_year for i in funding_year.all()),
                ", ".join(i.name for i in proj.funding_sources.all()),
                ", ".join(i.full_name for i in proj.DFO_project_authority.all()),
                proj.biological_process_type_1 if proj else None,
                proj.biological_process_type_2 if proj else None,
                ", ".join(i.name for i in proj.biological_process_type_3.all()),
                ", ".join(i.name for i in proj.activity_type_1.all()),
                ", ".join(i.name for i in proj.activity_type_2.all()),
                ", ".join(i.name for i in proj.activity_type_3.all()),
                proj.area if proj else None,
                obj.river.name if obj.river else None,
                obj.river.species if obj.river else None,
                sub_tmp,
                obj.river.stock_management_unit if obj.river else None,
                cui_tmp,
                du_tmp,
                obj.river.du_number if obj.river else None,
                cu_tmp,
                wtr_tmp,
                proj.first_nation.name if proj.first_nation else None,
                proj.funding_recipient if proj else None,
                proj.organization_program.name if proj.organization_program else None,
                ", ".join(i.name for i in proj.hatchery_name.all()),
                sample.unique_activity_outcome_number,
                sample.sampling_outcome,
                sample.outcome_delivered,
                sample.outcome_quality,
                sample.sample_outcome_comment,
                sample.date_last_modified,
                sample.last_modified_by.get_full_name() if sample.last_modified_by else None,
            ]
            for col_num, cell_value in enumerate(sample_rows, 1):
                cell = sample_outcome_worksheet.cell(row=samp_num, column=col_num)
                cell.value = cell_value

        report_outcomes = models.ReportOutcome.objects.filter(activity_outcome=obj.id)
        for report in report_outcomes:
            rep_num += 1
            report_rows = [
                proj.agreement_number if proj else None,
                proj.project_number if proj else None,
                proj.name if proj else None,
                ", ".join(i.funding_year for i in funding_year.all()),
                ", ".join(i.name for i in proj.funding_sources.all()),
                ", ".join(i.full_name for i in proj.DFO_project_authority.all()),
                proj.biological_process_type_1 if proj else None,
                proj.biological_process_type_2 if proj else None,
                ", ".join(i.name for i in proj.biological_process_type_3.all()),
                ", ".join(i.name for i in proj.activity_type_1.all()),
                ", ".join(i.name for i in proj.activity_type_2.all()),
                ", ".join(i.name for i in proj.activity_type_3.all()),
                proj.area if proj else None,
                obj.river.name if obj.river else None,
                obj.river.species if obj.river else None,
                sub_tmp,
                obj.river.stock_management_unit if obj.river else None,
                cui_tmp,
                du_tmp,
                obj.river.du_number if obj.river else None,
                cu_tmp,
                wtr_tmp,
                proj.first_nation.name if proj.first_nation else None,
                proj.funding_recipient if proj else None,
                proj.organization_program.name if proj.organization_program else None,
                ", ".join(i.name for i in proj.hatchery_name.all()),
                report.unique_activity_outcome_number,
                report.reporting_outcome,
                report.task_id,
                report.task_name,
                report.task_description,
                report.task_met,
                report.site,
                report.report_link.document_name if report.report_link else None,
                report.report_outcome_comment,
                report.reporting_outcome_metric,
                report.reporting_outcome_metric_unit,
                report.data_sharing,
                report.date_last_modified,
                report.last_modified_by.get_full_name() if report.last_modified_by else None,
            ]
            for col_num, cell_value in enumerate(report_rows, 1):
                cell = reporting_outcome_worksheet.cell(row=rep_num, column=col_num)
                cell.value = cell_value

    workbook.save(response)

    return response