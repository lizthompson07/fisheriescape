import csv
import itertools
import os
from datetime import datetime

from django.utils import timezone
import pytz
from bokeh.embed import components
from bokeh.layouts import column
from bokeh.models import Title
from bokeh.plotting import figure
from bokeh.resources import CDN
from django.db.models.functions import Concat
from django.urls import reverse
from openpyxl import load_workbook

from bio_diversity import models, utils
from bio_diversity.scripts import model_export
from bio_diversity.static.calculation_constants import in_out_dict, distribution_locc_list, collection_locc_list
from dm_apps import settings


class ExcelReport:
    # figure out the filename
    target_dir = os.path.join(settings.BASE_DIR, 'media', 'temp')
    target_file = "temp_export.xlsx"
    target_file_path = os.path.join(target_dir, target_file)
    if hasattr(settings, "MEDIA_ROOT"):  # when this is loaded on Azure cloud, there is no attribute call MEDIA_ROOT and upload/download is handled differently
        target_url = os.path.join(settings.MEDIA_ROOT, 'temp', target_file)
    else:
        target_url = os.path.join('temp', target_file)  # basically just a dummy name. will be handled differently when using azure cloud
    template_file_path = None
    wb = None

    template_dir = os.path.join(settings.BASE_DIR, 'bio_diversity', 'static', "report_templates")

    def load_wb(self, template_name):
        self.template_file_path = os.path.join(self.template_dir, template_name)
        self.wb = load_workbook(filename=self.template_file_path)

    def get_sheet(self, sheet_name):
        ws = None
        try:
            ws = self.wb[sheet_name]
        except KeyError:
            print(sheet_name, "is not a valid name of a worksheet")
        return ws

    def copy_template(self, sheet_name, template_name="template"):
        ws = self.wb[template_name]
        ws.title = sheet_name
        self.wb.copy_worksheet(ws).title = str(template_name)
        try:
            ws = self.wb[sheet_name]
        except KeyError:
            print(sheet_name, "is not a valid name of a worksheet")
        return ws

    def save_wb(self):
        self.wb.save(self.target_file_path)


def cont_treat_feed_writer(ws, cont_evnt_list, row_count, treat_row_count, feed_row_count, end_date=timezone.now()):
    treat_list = []
    feed_list = []
    for contx_dict in cont_evnt_list:
        evnt_id = contx_dict["evnt_id"]
        # cont_evnt = [evntc, date, direction, container]
        ws['A' + str(row_count)].value = evnt_id.start_datetime.date()
        ws['B' + str(row_count)].value = evnt_id.evntc_id.name
        ws['C' + str(row_count)].value = contx_dict["cont_id"].name
        ws['D' + str(row_count)].value = in_out_dict[contx_dict["destination"]]
        row_count += 1
        if contx_dict["destination"]:
            start_date = evnt_id.start_datetime
            treat_list.extend(contx_dict["cont_id"].cont_treatments(start_date, end_date))
            feed_list.extend(contx_dict["cont_id"].feed_history(start_date, end_date))
        if not contx_dict["destination"] and contx_dict is not None:
            end_date = evnt_id.start_datetime

    for treat in treat_list:
        ws['G' + str(treat_row_count)].value = treat.envtc_id.name
        ws['H' + str(treat_row_count)].value = treat.start_date
        ws['I' + str(treat_row_count)].value = treat.cont.__str__()
        ws['J' + str(treat_row_count)].value = treat.concentration_str
        ws['K' + str(treat_row_count)].value = "{} {}".format(treat.amt, treat.unit_id.name)
        ws['L' + str(treat_row_count)].value = treat.duration
        treat_row_count += 1

    for feed_contx in feed_list:
        feed_dict = feed_contx.feed_props
        ws['O' + str(feed_row_count)].value = feed_dict["str"]
        ws['P' + str(feed_row_count)].value = feed_contx.evnt_id.start_date
        ws['Q' + str(feed_row_count)].value = feed_contx.container.__str__()
        ws['R' + str(feed_row_count)].value = feed_dict["freq"]
        ws['S' + str(feed_row_count)].value = feed_dict["method"]
        ws['T' + str(feed_row_count)].value = feed_dict["comments"]
        feed_row_count += 1
    return row_count, treat_row_count, feed_row_count, end_date


def generate_facility_tank_report(request, facic_id, at_date=timezone.now()):
    report = ExcelReport()
    report.load_wb("facility_tank_template.xlsx")

    facic = models.FacilityCode.objects.filter(pk=facic_id).get()

    tank_qs = models.Tank.objects.filter(facic_id=facic).order_by('name')
    tray_qs = models.Tray.objects.filter(trof_id__facic_id=facic, end_date__isnull=True).order_by(Concat('trof_id__name', 'name'))

    draw_qs = models.Drawer.objects.filter(heat_id__facic_id=facic).order_by(Concat('heat_id__name', 'name'))
    cup_qs = models.Cup.objects.filter(draw_id__heat_id__facic_id=facic, end_date__isnull=True).order_by(
        Concat('draw_id__heat_id__name', 'draw_id__name', 'name'))

    qs_list = [("Tank", tank_qs, "tank"), ("Trough", tray_qs, "tray"), ("Drawer", draw_qs, "draw"), ("Cup", cup_qs, "cup")]

    # to order workshees so the first sheet comes before the template sheet, rename the template and then copy the
    # renamed sheet, then rename the copy to template so it exists for other sheets to be created from

    for sheet_name, qs, cont_code in qs_list:
        ws = report.copy_template(sheet_name)

        # start writing data at row 3 in the sheet
        row_count = 3
        for item in qs:

            ws['A' + str(row_count)].value = item.__str__()

            cnt = 0
            year_coll_set = set()
            indv_list, grp_list = item.fish_in_cont(select_fields=["indv_id__grp_id__stok_id",
                                                                   "indv_id__grp_id__coll_id"], at_date=at_date)
            if indv_list:
                ws['B' + str(row_count)].value = "Y"
                cnt += len(indv_list)
                year_coll_set |= set([indv.stok_year_coll_str() for indv in indv_list])
            if grp_list:
                for grp in grp_list:
                    cnt += grp.count_fish_in_group()
                    year_coll_set |= {grp.__str__()}
            ws['C' + str(row_count)].value = cnt
            ws['D' + str(row_count)].value = str(', '.join(set(year_coll_set)))

            feed_str = item.cont_feed(get_string=True)
            ws['E' + str(row_count)].value = feed_str
            cont_url =  "bio_diversity:details_" + cont_code
            ws['F' + str(row_count)].value = request.build_absolute_uri(reverse(cont_url, args=[item.id]))

            row_count += 1

    report.save_wb()

    return report.target_url


def generate_stock_code_report(stok_id, coll_id, year, start_date=utils.aware_min(), end_date=timezone.now()):
    # report is given a stock code and returns location of all associated fish
    report = ExcelReport()
    report.load_wb("stock_code_report_template.xlsx")

    start_date = utils.naive_to_aware(start_date)
    end_date = utils.naive_to_aware(end_date)

    indv_qs = models.Individual.objects.all()
    grp_qs = models.Group.objects.filter(grp_valid=True)

    if stok_id:
        indv_qs = indv_qs.filter(stok_id=stok_id)
        grp_qs = grp_qs.filter(stok_id=stok_id)
    if coll_id:
        indv_qs = indv_qs.filter(coll_id=coll_id)
        grp_qs = grp_qs.filter(coll_id=coll_id)
    if year:
        indv_qs = indv_qs.filter(indv_year=year)
        grp_qs = grp_qs.filter(grp_year=year)

    indv_qs = indv_qs.select_related("stok_id", "coll_id")
    grp_qs = grp_qs.select_related("stok_id", "coll_id")

    ws_indv = report.get_sheet("Individuals")
    ws_grp = report.get_sheet("Groups")

    ws_indv['A1'].value = "Stock: {}".format(stok_id.name)
    ws_grp['A1'].value = "Stock: {}".format(stok_id.name)
    # start writing data at row 3 in the sheet
    row_count = 3
    for item in indv_qs:
        ws_indv['A' + str(row_count)].value = item.pit_tag
        ws_indv['B' + str(row_count)].value = item.indv_year
        ws_indv['C' + str(row_count)].value = item.coll_id.name
        ws_indv['D' + str(row_count)].value = item.prog_group(get_string=True)
        ws_indv['E' + str(row_count)].value = ', '.join([cont.__str__() for cont in item.current_tank(end_date)])

        item_indvd = models.IndividualDet.objects.filter(indvd_valid=True, anidc_id__name="Animal Health",
                                                         adsc_id__isnull=False, anix_id__indv_id=item).select_related("adsc_id")

        indvd_str_list = [indvd.adsc_id.name for indvd in item_indvd]
        indvd_str = ", ".join(indvd_str_list)
        ws_indv['F' + str(row_count)].value = indvd_str

        item_sexd = item.individual_detail("Gender")
        if item_sexd:
            ws_indv['G' + str(row_count)].value = str(item_sexd)
        ws_indv['H' + str(row_count)].value = str(item.indv_valid)

        row_count += 1

    row_count = 3
    for item in grp_qs:
        current_cont_str = ', '.join([cont.__str__() for cont in item.current_cont(end_date)])
        ws_grp['A' + str(row_count)].value = item.pk
        ws_grp['B' + str(row_count)].value = item.grp_year
        ws_grp['C' + str(row_count)].value = item.coll_id.name
        ws_grp['D' + str(row_count)].value = item.prog_group(get_string=True)
        ws_grp['E' + str(row_count)].value = current_cont_str
        ws_grp['G' + str(row_count)].value = item.avg_len(at_date=end_date)
        ws_grp['H' + str(row_count)].value = item.avg_weight(at_date=end_date)
        ws_grp['I' + str(row_count)].value = item.count_fish_in_group(end_date)
        grp_history = item.get_cont_history(start_date=start_date, end_date=end_date)
        cont_str = ""
        for contx_dict in grp_history:
            if contx_dict["destination"] is not None:
                cont_str += contx_dict["cont_id"].name + ", "
        if not cont_str:
            cont_str = current_cont_str
        ws_grp['J' + str(row_count)].value = cont_str

        row_count += 1

    report.save_wb()

    return report.target_url


def generate_morts_report(request, facic_id=None, prog_id=None, stok_id=None, year=None, coll_id=None,
                          start_date=utils.aware_min(), end_date=timezone.now()):
    # report is given some filter criteria, returns all dead fish details.
    report = ExcelReport()
    report.load_wb("mortality_report_template.xlsx")
    start_date = utils.naive_to_aware(start_date)
    end_date = utils.naive_to_aware(end_date)

    mort_anidc = models.AnimalDetCode.objects.filter(name="Mortality Observation").get()

    # select all individuals
    indv_mortd_qs = models.IndividualDet.objects.filter(anidc_id=mort_anidc,
                                                        anix_id__evnt_id__start_datetime__gte=start_date,
                                                        anix_id__evnt_id__end_datetime__lte=end_date)
    grp_mortd_qs = models.GroupDet.objects.filter(anidc_id=mort_anidc,
                                                  anix_id__evnt_id__start_datetime__gte=start_date,
                                                  anix_id__evnt_id__end_datetime__lte=end_date)
    samp_mortd_qs = models.SampleDet.objects.filter(anidc_id=mort_anidc,
                                                    samp_id__anix_id__evnt_id__start_datetime__gte=start_date,
                                                    samp_id__anix_id__evnt_id__end_datetime__lte=end_date)
    if facic_id:
        indv_mortd_qs = indv_mortd_qs.filter(anix_id__evnt_id__facic_id=facic_id)
        grp_mortd_qs = grp_mortd_qs.filter(anix_id__evnt_id__facic_id=facic_id)
        samp_mortd_qs = samp_mortd_qs.filter(samp_id__anix_id__evnt_id__facic_id=facic_id)
    if prog_id:
        indv_mortd_qs = indv_mortd_qs.filter(anix_id__evnt_id__prog_id=prog_id)
        grp_mortd_qs = grp_mortd_qs.filter(anix_id__evnt_id__prog_id=prog_id)
        samp_mortd_qs = samp_mortd_qs.filter(samp_id__anix_id__evnt_id__prog_id=prog_id)
    if stok_id:
        indv_mortd_qs = indv_mortd_qs.filter(anix_id__indv_id__stok_id=stok_id)
        grp_mortd_qs = grp_mortd_qs.filter(anix_id__grp_id__stok_id=stok_id)
        samp_mortd_qs = samp_mortd_qs.filter(samp_id__anix_id__grp_id__stok_id=stok_id)
    if year:
        indv_mortd_qs = indv_mortd_qs.filter(anix_id__indv_id__indv_year=year)
        grp_mortd_qs = grp_mortd_qs.filter(anix_id__grp_id__grp_year=year)
        samp_mortd_qs = samp_mortd_qs.filter(samp_id__anix_id__grp_id__indv_year=year)
    if coll_id:
        indv_mortd_qs = indv_mortd_qs.filter(anix_id__indv_id__coll_id=coll_id)
        grp_mortd_qs = grp_mortd_qs.filter(anix_id__grp_id__coll_id=coll_id)
        samp_mortd_qs = samp_mortd_qs.filter(samp_id__anix_id__grp_id__coll_id=coll_id)

    indv_mortd_qs.select_related("anix_id", "anix_id__indv_id", "anix__id__indv_id__coll_id", "anix_id__indv_id__stok_id", "anix_id__evnt_id")
    grp_mortd_qs.select_related("anix_id", "anix_id__grp_id", "anix__id__grp_id__coll_id", "anix_id__grp_id__stok_id", "anix_id__evnt_id")
    samp_mortd_qs.select_related("anix_id", "anix_id__grp_id", "anix__id__grp_id__coll_id", "anix_id__grp_id__stok_id", "anix_id__evnt_id")

    # to order worksheets so the first sheet comes before the template sheet, rename the template and then copy the
    # renamed sheet, then rename the copy to template so it exists for other sheets to be created from
    ws_indv = report.get_sheet("Individuals")
    ws_samp = report.get_sheet("Samples")
    ws_grp = report.get_sheet("Groups")

    # start writing data at row 3 in the sheet
    row_count = 3
    for indvd in indv_mortd_qs:
        indv_id = indvd.anix_id.indv_id
        mort_date = utils.naive_to_aware(indvd.detail_date)
        ws_indv['A' + str(row_count)].value = indv_id.pit_tag
        ws_indv['B' + str(row_count)].value = indv_id.stok_id.name
        ws_indv['C' + str(row_count)].value = indv_id.indv_year
        ws_indv['D' + str(row_count)].value = indv_id.coll_id.name
        ws_indv['E' + str(row_count)].value = indv_id.prog_group(get_string=True)
        ws_indv['F' + str(row_count)].value = indv_id.current_cont(at_date=mort_date, valid_only=False, get_string=True)
        ws_indv['G' + str(row_count)].value = indv_id.individual_subj_detail("Gender", before_date=mort_date)
        ws_indv['H' + str(row_count)].value = mort_date
        ws_indv['I' + str(row_count)].value = indv_id.individual_detail("Length", before_date=mort_date)
        ws_indv['J' + str(row_count)].value = indv_id.individual_detail("Weight", before_date=mort_date)
        ws_indv['K' + str(row_count)].value = indv_id.individual_evnt_details(indvd.anix_id.evnt_id)
        ws_indv['L' + str(row_count)].value = request.build_absolute_uri(reverse("bio_diversity:details_indv",
                                                                                 args=[indv_id.id]))

        row_count += 1

    row_count = 3
    for sampd in samp_mortd_qs:
        samp_id = sampd.samp_id
        mort_date = utils.naive_to_aware(sampd.detail_date)
        ws_samp['A' + str(row_count)].value = samp_id.samp_num
        ws_samp['B' + str(row_count)].value = getattr(samp_id.stok_id, "name", "")
        ws_samp['C' + str(row_count)].value = samp_id.indv_year
        ws_samp['D' + str(row_count)].value = getattr(samp_id.coll_id, "name", "")
        ws_samp['E' + str(row_count)].value = samp_id.prog_group(get_string=True)
        ws_samp['F' + str(row_count)].value = samp_id.cont(get_string=True)
        ws_samp['G' + str(row_count)].value = samp_id.sample_detail("Gender", before_date=mort_date)
        ws_samp['H' + str(row_count)].value = mort_date
        ws_samp['I' + str(row_count)].value = samp_id.sample_detail("Length", before_date=mort_date)
        ws_samp['J' + str(row_count)].value = samp_id.sample_detail("Weight", before_date=mort_date)
        ws_samp['K' + str(row_count)].value = samp_id.comments
        ws_samp['L' + str(row_count)].value = request.build_absolute_uri(reverse("bio_diversity:details_samp",
                                                                                 args=[samp_id.id]))
        row_count += 1

    row_count = 3
    for grpd in grp_mortd_qs:
        grp_id = grpd.anix_id.anix_id.grp_id
        mort_date = utils.naive_to_aware(grpd.detail_date)
        ws_grp['A' + str(row_count)].value = grp_id.stok_id.name
        ws_grp['B' + str(row_count)].value = grp_id.grp_year
        ws_grp['C' + str(row_count)].value = grp_id.coll_id.name
        ws_grp['D' + str(row_count)].value = grp_id.prog_group(get_string=True)
        ws_grp['E' + str(row_count)].value = grp_id.current_cont(at_date=mort_date, valid_only=False, get_string=True)
        ws_grp['F' + str(row_count)].value = grpd.anix_id.samp_detail("Gender")
        ws_grp['G' + str(row_count)].value = mort_date
        ws_grp['H' + str(row_count)].value = grpd.anix_id.samp_detail("Length")
        ws_grp['I' + str(row_count)].value = grpd.anix_id.samp_detail("Weight")
        ws_grp['J' + str(row_count)].value = request.build_absolute_uri(reverse("bio_diversity:details_grp",
                                                                                 args=[grp_id.id]))
        row_count += 1

    report.save_wb()

    return report.target_url


def generate_detail_report(adsc_id, prog_id, stok_id=None):
    # report is given an animal detail subjective code (skinny/precocious) and returns
    # all fish with that detail
    # group and that detail count
    # container breakdown of the detail

    report = ExcelReport()
    report.load_wb("detail_report_template.xlsx")

    indvd_set = models.IndividualDet.objects.all().select_related("anix_id__indv_id__stok_id", "anix_id__evnt_id__prog_id")
    if stok_id:
        indvd_set = indvd_set.filter(anix_id__indv_id__stok_id=stok_id)
    if prog_id:
        indvd_set = indvd_set.filter(anix_id__evnt_id__prog_id=prog_id)
    indvd_set = indvd_set.filter(adsc_id=adsc_id, anix_id__indv_id__isnull=False). \
        select_related("anix_id__indv_id", "anix_id__indv_id__stok_id", "anix_id__indv_id__coll_id", )
    indv_list = list(dict.fromkeys([indvd.anix_id.indv_id for indvd in indvd_set]))
    sampd_set = models.SampleDet.objects.filter(adsc_id=adsc_id, samp_id__anix_id__grp_id__isnull=False). \
        select_related("samp_id__anix_id__grp_id", "samp_id__anix_id__grp_id__stok_id", "samp_id__anix_id__grp_id__coll_id", )
    grp_list = list(dict.fromkeys([sampd.samp_id.anix_id.grp_id for sampd in sampd_set]))

    # to order workshees so the first sheet comes before the template sheet, rename the template and then copy the
    # renamed sheet, then rename the copy to template so it exists for other sheets to be created from
    ws_indv = report.get_sheet('Individuals')
    ws_grp = report.get_sheet('Groups')

    ws_indv['A1'].value = "Detail: {}".format(adsc_id.name)
    # start writing data at row 3 in the sheet
    row_count = 3
    for item in indv_list:
        ws_indv['A' + str(row_count)].value = item.pit_tag
        ws_indv['B' + str(row_count)].value = item.stok_id.name
        ws_indv['C' + str(row_count)].value = item.indv_year
        ws_indv['D' + str(row_count)].value = item.coll_id.name
        ws_indv['E' + str(row_count)].value = ', '.join([cont.__str__() for cont in item.current_tank()])
        ws_indv['F' + str(row_count)].value = item.individual_detail("Gender")
        ws_indv['G' + str(row_count)].value = item.indv_valid

        indvd_qs = models.IndividualDet.objects.filter(adsc_id=adsc_id, anix_id__indv_id=item).order_by("-detail_date")
        ws_indv['H' + str(row_count)].value = len(indvd_qs)
        ws_indv['I' + str(row_count)].value = indvd_qs.first().detail_date
        ws_indv['J' + str(row_count)].value = indvd_qs.last().detail_date
        ws_indv['K' + str(row_count)].value = indvd_qs.first().anix_id.evnt_id.__str__()
        ws_indv['L' + str(row_count)].value = indvd_qs.last().anix_id.evnt_id.__str__()

        row_count += 1

    ws_grp['A1'].value = "Detail: {}".format(adsc_id.name)
    # start writing data at row 3 in the sheet
    row_count = 3
    for item in grp_list:
        ws_grp['A' + str(row_count)].value = item.stok_id.name
        ws_grp['B' + str(row_count)].value = item.grp_year
        ws_grp['C' + str(row_count)].value = item.coll_id.name
        ws_grp['D' + str(row_count)].value = ', '.join([cont.__str__() for cont in item.current_tank()])
        ws_grp['E' + str(row_count)].value = item.grp_valid

        sampd_qs = models.SampleDet.objects.filter(adsc_id=adsc_id, samp_id__anix_id__grp_id=item).order_by("-detail_date")
        ws_grp['F' + str(row_count)].value = len(sampd_qs)
        ws_grp['G' + str(row_count)].value = sampd_qs.first().detail_date
        ws_grp['H' + str(row_count)].value = sampd_qs.last().detail_date
        ws_grp['I' + str(row_count)].value = sampd_qs.first().samp_id.anix_id.evnt_id.__str__()
        ws_grp['J' + str(row_count)].value = sampd_qs.last().samp_id.anix_id.evnt_id.__str__()
        row_count += 1

    report.save_wb()
    return report.target_url


def write_location_header(ws, row_count, rive_name, site_name, site_location, coll_str, grp_str, request):
    ws['A' + str(row_count)].value = rive_name
    ws['B' + str(row_count)].value = site_name
    ws['C' + str(row_count)].value = site_location.start_date
    ws['D' + str(row_count)].value = site_location.locc_id.name
    ws['E' + str(row_count)].value = coll_str
    ws['F' + str(row_count)].value = grp_str
    ws['G' + str(row_count)].value = site_location.comments
    ws['H' + str(row_count)].value = request.build_absolute_uri(
        reverse("bio_diversity:details_loc", args=[site_location.id]))


def write_location_to_sheets(ws, site_location, row_count, rive_name, site_name, request, dist_ws, dist_row_count, coll_ws, coll_row_count):
    grps = [(anix.grp_id.__str__(), anix.grp_id.prog_group(get_string=True)) for anix in
            site_location.animal_details.filter(grp_id__isnull=False).select_related("grp_id__stok_id",
                                                                                     "grp_id__coll_id")]
    indvs = ["{}-{}-{}".format(anix_tup[0], anix_tup[1], anix_tup[2])
             for anix_tup in
             site_location.animal_details.filter(indv_id__isnull=False).values_list("indv_id__stok_id__name",
                                                                                    "indv_id__indv_year",
                                                                                    "indv_id__coll_id__name", ).distinct()]
    loc_cnt_qs = models.Count.objects.filter(loc_id_id=site_location.pk).select_related("cntc_id", "coll_id")
    locd_qs = models.LocationDet.objects.filter(loc_id_id=site_location.pk).select_related("locdc_id", "ldsc_id")

    if len(grps) >= 1:
        coll_str = grps[0][0]
        grp_str = grps[0][1]
    elif len(indvs) >= 1:
        coll_str = indvs[0]
        grp_str = ""
    elif len(loc_cnt_qs) >= 1:
        coll_str = loc_cnt_qs.first().coll_str
        grp_str = loc_cnt_qs.first().prog_group(get_string=True)
    else:
        coll_str = ""
        grp_str = ""

    write_location_header(ws, row_count, rive_name, site_name, site_location, coll_str, grp_str, request)

    if site_location.locc_id.name in distribution_locc_list:
        relm = ", ".join(locd_qs.filter(locdc_id__name="Release Method").values_list('ldsc_id__name', flat=True))
        truck_list = locd_qs.filter(locdc_id__name="Truck").values_list('det_val', flat=True)
        if len(truck_list) == 1:
            truck = truck_list[0]
        else:
            truck = ", ".join(truck_list)

        acclimation_qs = locd_qs.filter(locdc_id__name="Acclimation Time").values_list('det_val', flat=True)
        if len(acclimation_qs) == 1:
            acclimation_time = acclimation_qs[0]
        else:
            acclimation_time = ", ".join(acclimation_qs)

        for cnt in loc_cnt_qs:
            write_location_header(dist_ws, dist_row_count, rive_name, site_name, site_location, coll_str, grp_str, request)
            dist_ws['I' + str(dist_row_count)].value = relm
            dist_ws['J' + str(dist_row_count)].value = cnt.cntc_id.name
            dist_ws['K' + str(dist_row_count)].value = cnt.cnt
            dist_ws['L' + str(dist_row_count)].value = cnt.cnt_detail()
            dist_ws['M' + str(dist_row_count)].value = cnt.cnt_detail("Weight")
            dist_ws['N' + str(dist_row_count)].value = cnt.cnt_detail("Lifestage")
            dist_ws['O' + str(dist_row_count)].value = truck
            dist_ws['Q' + str(dist_row_count)].value = acclimation_time

            dist_row_count += 1

    if site_location.locc_id.name in collection_locc_list:
        for cnt in loc_cnt_qs:
            write_location_header(coll_ws, coll_row_count, rive_name, site_name, site_location, coll_str, grp_str,
                                  request)
            coll_ws['I' + str(coll_row_count)].value = cnt.cntc_id.name
            coll_ws['J' + str(coll_row_count)].value = cnt.cnt
            coll_ws['K' + str(coll_row_count)].value = cnt.cnt_detail()
            coll_ws['L' + str(coll_row_count)].value = cnt.cnt_detail("Weight")
            coll_ws['M' + str(coll_row_count)].value = cnt.cnt_detail("Lifestage")

            coll_row_count += 1

    row_count += 1
    return row_count, dist_row_count, coll_row_count


def generate_sites_report(sites_list, locations_list, start_date=None, end_date=None, request=None):
    report = ExcelReport()
    report.load_wb("site_report_template.xlsx")

    ws = report.get_sheet("Sites")
    ws_indv = report.get_sheet("Individuals")
    ws_dist = report.get_sheet("Distributions")
    ws_coll = report.get_sheet("Collections")

    if not start_date:
        start_date = "Not Picked"
        ws['B1'].value = start_date
        ws_indv['B1'].value = start_date
        start_date = utils.aware_min()
    else:
        ws['B1'].value = start_date
        ws_indv['B1'].value = start_date

    if not end_date:
        end_date = timezone.now()
    ws['B2'].value = end_date
    ws_indv['B2'].value = end_date

    # start writing data at row 4 in the sheet
    row_count = 4
    dist_row_count = 4
    coll_row_count = 4
    # split off locations with no sites
    no_sites_list = [location for location in locations_list if not location.relc_id]

    site_locations = models.Location.objects.filter(relc_id__in=sites_list, loc_date__lte=end_date,
                                                    loc_date__gte=start_date)

    locations_list = [location for location in site_locations]
    no_locs_list = []
    for site in sites_list:
        site_name = site.name
        rive_name = site.rive_id.name
        site_locations = [location for location in locations_list if location.relc_id.pk == site.pk]
        if not site_locations:
            no_locs_list.append(site)

        for site_location in site_locations:
            row_count, dist_row_count, coll_row_count = write_location_to_sheets(ws, site_location, row_count,
                                                                                 rive_name, site_name, request,
                                                                                 ws_dist, dist_row_count, ws_coll,
                                                                                 coll_row_count)

    for location in no_sites_list:
        if location.rive_id:
            rive_name = location.rive_id.name
        else:
            rive_name = None
        row_count, dist_row_count, coll_row_count = write_location_to_sheets(ws, location, row_count, rive_name,
                                                                             None, request, ws_dist, dist_row_count,
                                                                             ws_coll, coll_row_count)

    for site in no_locs_list:
        ws['A' + str(row_count)].value = site.rive_id.name
        ws['B' + str(row_count)].value = site.name
        ws['D' + str(row_count)].value = "No Events at location"
        row_count += 1

    anix_indv_set = models.AniDetailXref.objects.filter(loc_id__in=locations_list, indv_id__isnull=False) \
        .select_related("indv_id", "indv_id__coll_id", "indv_id__stok_id", "loc_id", "loc_id__locc_id", "loc_id__relc_id")
    indv_list = [(anix.indv_id, anix.loc_id) for anix in anix_indv_set]
    row_count = 4
    for indv, loc in indv_list:
        ws_indv['A' + str(row_count)].value = indv.pit_tag
        ws_indv['B' + str(row_count)].value = indv.stok_id.name
        ws_indv['C' + str(row_count)].value = indv.indv_year
        ws_indv['D' + str(row_count)].value = indv.coll_id.name
        ws_indv['E' + str(row_count)].value = indv.individual_detail(anidc_name="Lifestage", before_date=loc.loc_date)
        ws_indv['F' + str(row_count)].value = loc.relc_id.name
        ws_indv['G' + str(row_count)].value = loc.loc_date
        ws_indv['H' + str(row_count)].value = loc.locc_id.name
        ws_indv['I' + str(row_count)].value = indv.prog_group(get_string=True)
        ws_indv['J' + str(row_count)].value = indv.individual_detail(anidc_name="Gender", before_date=loc.loc_date)
        ws_indv['K' + str(row_count)].value = indv.individual_detail(anidc_name="Length", before_date=loc.loc_date)
        ws_indv['L' + str(row_count)].value = indv.individual_detail(anidc_name="Weight", before_date=loc.loc_date)
        row_count += 1
    report.save_wb()

    return report.target_url


def generate_individual_report(indv_id):
    report = ExcelReport()
    report.load_wb("individual_report_template.xlsx")

    ws_evnt = report.get_sheet('Event History')
    ws_hist = report.get_sheet('Heritage')
    ws_cont = report.get_sheet('Containers')
    ws_dets = report.get_sheet('Details')

    # -----------------Heritage Sheet---------------
    prnt_grp_set = indv_id.get_parent_history()
    row_count = 5
    true_false_dict = {True: "Yes", False: "No"}
    for grp_tuple in prnt_grp_set:
        grp_id = grp_tuple[1]
        if grp_id:
            conts = ', '.join([cont.__str__() for cont in grp_id.current_cont()])
            ws_hist['A' + str(row_count)].value = grp_tuple[2]
            ws_hist['B' + str(row_count)].value = grp_tuple[0]
            ws_hist['C' + str(row_count)].value = grp_id.__str__()
            ws_hist['D' + str(row_count)].value = true_false_dict[grp_id.grp_valid]
            ws_hist['E' + str(row_count)].value = conts
            row_count += 1

    # -----------Events Sheet------------------
    # put in start and end dates
    ws_evnt['B2'].value = datetime.today().date()
    ws_evnt['B3'].value = indv_id.animal_details.first().evnt_id.facic_id.name
    ws_evnt['E2'].value = indv_id.__str__()
    ws_evnt['E3'].value = indv_id.pit_tag

    anix_evnt_set = indv_id.animal_details.filter(contx_id__isnull=True, loc_id__isnull=True, pair_id__isnull=True) \
        .order_by("-evnt_id__start_datetime").select_related('evnt_id', 'evnt_id__evntc_id', 'evnt_id__facic_id',
                                                             'evnt_id__prog_id', 'evnt_id__perc_id')
    evnt_list = list(dict.fromkeys([anix.evnt_id for anix in anix_evnt_set]))

    row_count = 6
    for evnt in evnt_list:
        ws_evnt['A' + str(row_count)].value = evnt.start_date
        ws_evnt['B' + str(row_count)].value = evnt.evntc_id.name
        ws_evnt['C' + str(row_count)].value = ""
        ws_evnt['D' + str(row_count)].value = indv_id.current_cont(at_date=utils.naive_to_aware(evnt.start_date), get_string=True)
        ws_evnt['E' + str(row_count)].value = evnt.comments

        row_count += 1

    for grp_tuple in prnt_grp_set:
        grp_id = grp_tuple[1]
        if grp_id:
            start_date = utils.naive_to_aware(grp_id.start_date())
            end_date = utils.naive_to_aware(grp_tuple[2])
            cont_evnt_list = grp_id.get_cont_history(start_date=start_date, end_date=end_date)
            for contx_dict in cont_evnt_list:
                evnt = contx_dict["evnt_id"]
                ws_evnt['A' + str(row_count)].value = evnt.start_date
                ws_evnt['B' + str(row_count)].value = evnt.evntc_id.name
                ws_evnt['C' + str(row_count)].value = grp_id.__str__()
                ws_evnt['D' + str(row_count)].value = grp_id.current_cont(at_date=utils.naive_to_aware(evnt.start_date))[0].name
                ws_evnt['E' + str(row_count)].value = evnt.comments
                row_count += 1

    # -----------------Container Sheet------------------------
    anix_evnt_set = indv_id.animal_details.filter(contx_id__isnull=False, loc_id__isnull=True, pair_id__isnull=True) \
        .order_by("-evnt_id__start_datetime", "-final_contx_flag") \
        .select_related('contx_id', 'contx_id__evnt_id__evntc_id', 'contx_id__evnt_id')
    contx_tuple_set = list(dict.fromkeys([(anix.contx_id, anix.final_contx_flag) for anix in anix_evnt_set]))
    cont_evnt_list = [utils.get_cont_evnt(contx) for contx in contx_tuple_set]
    row_count = 5
    row_count, treat_row_count, treat_end_date, feed_row_count = cont_treat_feed_writer(ws_cont, cont_evnt_list, row_count, row_count, row_count)

    for grp_tuple in prnt_grp_set:
        grp_id = grp_tuple[1]
        if grp_id:
            start_date = utils.naive_to_aware(grp_id.start_date())
            end_date = utils.naive_to_aware(grp_tuple[2])
            cont_evnt_list = grp_id.get_cont_history(start_date=start_date, end_date=end_date)
            treat_end_date = cont_treat_feed_writer(ws_cont, cont_evnt_list, row_count, treat_row_count, feed_row_count, end_date=treat_end_date)[2]

    # -----------------Details Sheet------------------------
    indvd_set = models.IndividualDet.objects.filter(anix_id__indv_id=indv_id).distinct(). \
        order_by("anidc_id__name", "adsc_id", "-detail_date").select_related("anidc_id", "adsc_id", )
    row_count = 5
    for indvd in indvd_set:
        adsc_name = ""
        if indvd.adsc_id:
            adsc_name = indvd.adsc_id.name
        ws_dets['A' + str(row_count)].value = indvd.detail_date
        ws_dets['B' + str(row_count)].value = indvd.anidc_id.name
        ws_dets['C' + str(row_count)].value = adsc_name
        ws_dets['D' + str(row_count)].value = indvd.det_val
        ws_dets['E' + str(row_count)].value = indv_id.current_cont(indvd.detail_date, get_string=True)
        ws_dets['F' + str(row_count)].value = indvd.comments
        row_count += 1

    indvt_set = models.IndTreatment.objects.filter(anix_id__indv_id=indv_id).distinct().select_related("indvtc_id",
                                                                                                       "unit_id")
    row_count = 5
    for indvt in indvt_set:
        ws_dets['I' + str(row_count)].value = indvt.start_date
        ws_dets['J' + str(row_count)].value = indvt.indvtc_id.name
        ws_dets['K' + str(row_count)].value = indvt.dose
        ws_dets['L' + str(row_count)].value = indvt.unit_id.name
        row_count += 1

    report.save_wb()

    return report.target_url


def generate_grp_report(grp_id):
    report = ExcelReport()
    report.load_wb("group_report_template.xlsx")

    ws_evnt = report.get_sheet('Event History')
    ws_hist = report.get_sheet('Heritage')
    ws_cont = report.get_sheet('Containers')

    # -----------------Heritage Sheet---------------
    prnt_grp_set = grp_id.get_parent_history()
    row_count = 5
    true_false_dict = {True: "Yes", False: "No"}
    for grp_tuple in prnt_grp_set:
        grp_id = grp_tuple[1]
        if grp_id:
            conts = ', '.join([cont.__str__() for cont in grp_id.current_cont()])
            ws_hist['A' + str(row_count)].value = grp_tuple[2]
            ws_hist['B' + str(row_count)].value = grp_tuple[0]
            ws_hist['C' + str(row_count)].value = grp_id.__str__()
            ws_hist['D' + str(row_count)].value = true_false_dict[grp_id.grp_valid]
            ws_hist['E' + str(row_count)].value = conts
            row_count += 1

    # -----------Events Sheet------------------
    # put in start and end dates
    ws_evnt['B2'].value = datetime.today().date()
    ws_evnt['B3'].value = grp_id.animal_details.first().evnt_id.facic_id.name
    ws_evnt['E2'].value = grp_id.__str__()
    ws_evnt['E3'].value = grp_id.prog_group(get_string=True)

    anix_evnt_set = grp_id.animal_details.filter(contx_id__isnull=True, loc_id__isnull=True, pair_id__isnull=True) \
        .order_by("-evnt_id__start_datetime").select_related('evnt_id', 'evnt_id__evntc_id', 'evnt_id__facic_id',
                                                             'evnt_id__prog_id', 'evnt_id__perc_id')
    evnt_list = list(dict.fromkeys([anix.evnt_id for anix in anix_evnt_set]))

    row_count = 6
    for evnt in evnt_list:
        ws_evnt['A' + str(row_count)].value = evnt.start_date
        ws_evnt['B' + str(row_count)].value = evnt.evntc_id.name
        ws_evnt['C' + str(row_count)].value = ""
        ws_evnt['D' + str(row_count)].value = grp_id.current_cont(at_date=utils.naive_to_aware(evnt.start_date), get_string=True)
        ws_evnt['E' + str(row_count)].value = evnt.comments
        row_count += 1

    for grp_tuple in prnt_grp_set:
        grp_id = grp_tuple[1]
        if grp_id:
            start_date = utils.naive_to_aware(grp_id.start_date())
            end_date = utils.naive_to_aware(grp_tuple[2])
            anix_evnt_set = grp_id.animal_details.filter(contx_id__isnull=True, loc_id__isnull=True,
                                                         pair_id__isnull=True, evnt_id__start_datetime__lte=end_date,
                                                         evnt_id__start_datetime__gte=start_date) \
                .order_by("-evnt_id__start_datetime").select_related('evnt_id', 'evnt_id__evntc_id', 'evnt_id__facic_id',
                                                                     'evnt_id__prog_id', 'evnt_id__perc_id')
            evnt_list = list(dict.fromkeys([anix.evnt_id for anix in anix_evnt_set]))
            for evnt in evnt_list:
                ws_evnt['A' + str(row_count)].value = evnt.start_date
                ws_evnt['B' + str(row_count)].value = evnt.evntc_id.name
                ws_evnt['C' + str(row_count)].value = grp_id.__str__()
                ws_evnt['D' + str(row_count)].value = grp_id.current_cont(at_date=utils.naive_to_aware(evnt.start_date))[0].name
                ws_evnt['E' + str(row_count)].value = evnt.comments
                row_count += 1

    # -----------------Container Sheet------------------------
    cont_evnt_list = grp_id.get_cont_history()
    row_count = 5
    row_count, treat_row_count, treat_end_date, feed_row_count = cont_treat_feed_writer(ws_cont, cont_evnt_list, row_count, row_count, row_count)

    for grp_tuple in prnt_grp_set:
        grp_id = grp_tuple[1]
        if grp_id:
            start_date = utils.naive_to_aware(grp_id.start_date())
            end_date = utils.naive_to_aware(grp_tuple[2])
            cont_evnt_list = grp_id.get_cont_history(end_date=end_date, start_date=start_date)
            treat_end_date = cont_treat_feed_writer(ws_cont, cont_evnt_list, row_count, treat_row_count, feed_row_count)[2]

    report.save_wb()

    return report.target_url


def generate_system_code_report():
    report = ExcelReport()
    report.load_wb("system_code_report_template.xlsx")
    ws = report.get_sheet('System Codes')
    model_list = model_export

    col_count = 0
    for model_type in model_list:
        col = utils.col_count_to_excel(col_count)
        ws[col + "2"].value = model_type.__name__
        code_qs = model_type.objects.all()
        row_count = 3
        for code in code_qs:
            ws[col + str(row_count)].value = code.__str__()
            row_count += 1
        col_count += 1

    report.save_wb()

    return report.target_url


def generate_samples_report(request, prog_id, facic_id, stok_id, coll_id, year, start_date, end_date):
    report = ExcelReport()
    report.load_wb("samples_report_template.xlsx")
    ws_vials = report.get_sheet('Vials')
    ws_tissue = report.get_sheet('Tissue Samples')
    ws_scale = report.get_sheet('Scale Envelopes')

    fill_samples_sheet(request, report, ws_vials, "Vial", prog_id, facic_id, stok_id, coll_id, year, start_date, end_date)
    fill_samples_sheet(request, report, ws_tissue, "Tissue Sample", prog_id, facic_id, stok_id, coll_id, year, start_date, end_date)
    fill_samples_sheet(request, report, ws_scale, "Scale Envelope", prog_id, facic_id, stok_id, coll_id, year, start_date, end_date)

    return report.target_url


def generate_events_report(request, prog_id, facic_id, start_date, end_date):
    report = ExcelReport()
    report.load_wb("events_report_template.xlsx")
    ws = report.get_sheet('Events')

    evnt_qs = models.Event.objects.all()
    if prog_id:
        evnt_qs = evnt_qs.filter(prog_id=prog_id)
    if facic_id:
        evnt_qs = evnt_qs.filter(facic_id=facic_id)
    if start_date:
        evnt_qs = evnt_qs.filter(start_datetime__gte=start_date)
    if end_date:
        evnt_qs = evnt_qs.filter(start_datetime__lte=end_date)

    evnt_qs = evnt_qs.select_related("evntc_id", "prog_id", "facic_id")

    row_count = 3
    for evnt_id in evnt_qs:
        ws["A" + str(row_count)].value = evnt_id.evntc_id.__str__()
        ws["B" + str(row_count)].value = evnt_id.start_datetime
        ws["C" + str(row_count)].value = evnt_id.end_datetime
        ws["D" + str(row_count)].value = evnt_id.prog_id.__str__()
        ws["E" + str(row_count)].value = evnt_id.facic_id.__str__()
        ws["F" + str(row_count)].value = evnt_id.created_date
        ws["G" + str(row_count)].value = evnt_id.perc_id.__str__()

        indv_cnt = models.AniDetailXref.objects.filter(evnt_id=evnt_id, indv_id__isnull=False).values("indv_id").distinct().count()
        grp_cnt = models.AniDetailXref.objects.filter(evnt_id=evnt_id, grp_id__isnull=False).values("indv_id").distinct().count()
        ws["H" + str(row_count)].value = indv_cnt
        ws["I" + str(row_count)].value = grp_cnt
        ws["J" + str(row_count)].value = evnt_id.comments
        ws["K" + str(row_count)].value = request.build_absolute_uri(reverse("bio_diversity:details_evnt", args=[evnt_id.id]))
        row_count += 1

    report.save_wb()

    return report.target_url


def generate_growth_chart(plot_fish):
    if type(plot_fish) == models.Individual:
        len_dets = models.IndividualDet.objects.filter(anidc_id__name="Length").filter(anix_id__indv_id=plot_fish)
        weight_dets = models.IndividualDet.objects.filter(anidc_id__name="Weight").filter(anix_id__indv_id=plot_fish)
    else:
        indv_list, grp_list = plot_fish.fish_in_cont(select_fields=[])

        # indvds
        indv_lens = models.IndividualDet.objects.filter(anidc_id__name="Length").filter(anix_id__indv_id__in=indv_list)
        indv_weights = models.IndividualDet.objects.filter(anidc_id__name="Weight").filter(anix_id__indv_id__in=indv_list)

        # sampds
        samp_lens = models.SampleDet.objects.filter(anidc_id__name="Length").filter(samp_id__anix_id__grp_id__in=grp_list)
        samp_weights = models.SampleDet.objects.filter(anidc_id__name="Weight").filter(samp_id__anix_id__grp_id__in=grp_list)

        # this is okay because both indvds and sampds have detail_dates and det_vals
        len_dets = list(itertools.chain(indv_lens, samp_lens))
        weight_dets = list(itertools.chain(indv_weights, samp_weights))

    x_len_data = []
    y_len_data = []
    for len_det in len_dets:
        x_len_data.append(datetime.combine(len_det.detail_date, utils.aware_min().time()))
        y_len_data.append(len_det.det_val)

    x_weight_data = []
    y_weight_data = []
    for weight_det in weight_dets:
        x_weight_data.append(datetime.combine(weight_det.detail_date, utils.aware_min().time()))
        y_weight_data.append(weight_det.det_val)

    x_cond_data = []
    y_cond_data = []
    if type(plot_fish) == models.Individual:
        for len_det in len_dets:
            weight_det = weight_dets.filter(detail_date=len_det.detail_date).first()
            if weight_det:
                x_cond_data.append(datetime.combine(len_det.detail_date, utils.aware_min().time()))
                y_cond_data.append(utils.condition_factor(len_det.det_val, weight_det.det_val))

    # create a new plot
    title_eng = "Growth Chart for fish"

    if x_cond_data:
        p_cond = figure(
            tools="pan,box_zoom,wheel_zoom,reset,save",
            x_axis_type='datetime',
            x_axis_label='Date',
            y_axis_label='Condition Factor',
            plot_width=600, plot_height=300,
        )
        p_cond.axis.axis_label_text_font_style = 'normal'
        p_cond.x(x=x_cond_data, y=y_cond_data, size=10)

    p_len = figure(
        tools="pan,box_zoom,wheel_zoom,reset,save",
        x_axis_type='datetime',
        x_axis_label='Date',
        y_axis_label='Length',
        plot_width=600, plot_height=300,
    )
    p_weight = figure(
        tools="pan,box_zoom,wheel_zoom,reset,save",
        x_axis_type='datetime',
        x_axis_label='Date',
        y_axis_label='Weight',
        plot_width=600, plot_height=300,
    )

    p_len.axis.axis_label_text_font_style = 'normal'
    p_weight.axis.axis_label_text_font_style = 'normal'
    p_len.add_layout(Title(text=title_eng, text_font_size="16pt"), 'above')
    p_len.x(x=x_len_data, y=y_len_data, size=10)
    p_weight.x(x=x_weight_data, y=y_weight_data, size=10)

    # ------------------------Data File------------------------------
    target_dir = os.path.join(settings.BASE_DIR, 'media', 'temp')
    target_file = "temp_export.csv"
    target_file_path = os.path.join(target_dir, target_file)
    target_url = os.path.join(settings.MEDIA_ROOT, 'temp', target_file)
    with open(target_file_path, 'w') as data_file:
        writer = csv.writer(data_file)
        writer.writerow(["Growth information for Fish {}".format(plot_fish.__str__())])
        writer.writerow(["Date", " Length (cm)", " Date", " Weight (g)"])
        writer.writerows(itertools.zip_longest(x_len_data, y_len_data, x_weight_data, y_weight_data))
    if x_cond_data:
        scirpt, div = components(column(p_len, p_weight, p_cond), CDN)
    else:
        scirpt, div = components(column(p_len, p_weight), CDN)
    return scirpt, div, target_url


def generate_maturity_rate(cont):
    hist_dict = {"Immature": 0, "Male": 0, "Female": 0}
    pit_tag_list = []
    gender_list = []
    indv_list, grp_list = cont.fish_in_cont(select_fields=[])

    indvd_set = models.IndividualDet.objects.filter(anidc_id__name="Gender", indvd_valid=True, anix_id__indv_id__in=indv_list).select_related(
        "anix_id__indv_id")

    for indvd in indvd_set:
        pit_tag_list.append(indvd.anix_id.indv_id.pit_tag)
        gender = indvd.det_val
        hist_dict[gender] += 1
        gender_list.append(gender)

    for indv in indv_list:
        if indv.pit_tag not in pit_tag_list:
            pit_tag_list.append(indv.pit_tag)
            gender_list.append("Unknown")

    genders = ["Immature", "Male", "Female"]

    # create a new plot
    title_eng = "Maturity Rate"

    p = figure(tools="pan,box_zoom,wheel_zoom,reset,save",
               x_range=genders,
               plot_width=600,
               plot_height=300,
               title="Maturity Rate")

    p.vbar(x=genders, top=[hist_dict["Immature"], hist_dict["Male"], hist_dict["Female"]], width=0.9)
    p.axis.axis_label_text_font_style = 'normal'
    p.add_layout(Title(text=title_eng, text_font_size="16pt"), 'above')

    p.xgrid.grid_line_color = None
    p.y_range.start = 0

    # ------------------------Data File------------------------------
    target_dir = os.path.join(settings.BASE_DIR, 'media', 'temp')
    target_file = "temp_export.csv"
    target_file_path = os.path.join(target_dir, target_file)
    target_url = os.path.join(settings.MEDIA_ROOT, 'temp', target_file)
    with open(target_file_path, 'w') as data_file:
        writer = csv.writer(data_file)
        writer.writerow(["Pit Tag", "Gender"])
        writer.writerows(itertools.zip_longest(pit_tag_list, gender_list))
    scirpt, div = components(p, CDN)
    return scirpt, div, target_url


def plot_date_data(x_data, y_data, y_label, title):
    # create a new plot
    title_eng = title

    p = figure(
        tools="pan,box_zoom,wheel_zoom,reset,save",
        x_axis_type='datetime',
        x_axis_label='Date',
        y_axis_label=y_label,
        plot_width=600, plot_height=400,
    )

    p.axis.axis_label_text_font_style = 'normal'
    p.add_layout(Title(text=title_eng, text_font_size="16pt"), 'above')
    p.line(x=x_data, y=y_data, line_width=3)

    # ------------------------Data File------------------------------
    target_dir = os.path.join(settings.BASE_DIR, 'media', 'temp')
    target_file = "temp_export.csv"
    target_file_path = os.path.join(target_dir, target_file)
    target_url = os.path.join(settings.MEDIA_ROOT, 'temp', target_file)
    with open(target_file_path, 'w') as data_file:
        writer = csv.writer(data_file)
        writer.writerow(["Date", y_label])
        writer.writerows(itertools.zip_longest(x_data, y_data))
    scirpt, div = components(p, CDN)
    return scirpt, div, target_url


def fill_samples_sheet(request, report, ws, anidc_name, prog_id, facic_id, stok_id, coll_id, year, start_date, end_date):
    anidc_id = models.AnimalDetCode.objects.filter(name=anidc_name).get()

    indvd_qs = models.IndividualDet.objects.filter(anidc_id=anidc_id,
                                                   anix_id__evnt_id__start_datetime__lte=end_date,
                                                   anix_id__evnt_id__start_datetime__gte=start_date).select_related(
        "anix_id__evnt_id", "anix_id__indv_id__stok_id", "anix_id__indv_id__coll_id", "anix_id__indv_id")
    sampd_qs = models.SampleDet.objects.filter(anidc_id=anidc_id,
                                               samp_id__anix_id__evnt_id__start_datetime__lte=end_date,
                                               samp_id__anix_id__evnt_id__start_datetime__gte=start_date).select_related(
        "samp_id__anix_id__evnt_id", "samp_id", "samp_id__anix_id__grp_id", "samp_id__anix_id__grp_id__coll_id",
        "samp_id__anix_id__grp_id__stok_id")

    if stok_id:
        indvd_qs = indvd_qs.filter(anix_id__indv_id__stok_id=stok_id)
        sampd_qs = sampd_qs.filter(samp_id__anix_id__grp_id__stok_id=stok_id)
    if facic_id:
        indvd_qs = indvd_qs.filter(anix_id__evnt_id__facic_id=facic_id)
        sampd_qs = sampd_qs.filter(samp_id__anix_id__evnt_id__facic_id=facic_id)
    if coll_id:
        indvd_qs = indvd_qs.filter(anix_id__indv_id__coll_id=coll_id)
        sampd_qs = sampd_qs.filter(samp_id__anix_id__grp_id__coll_id=coll_id)
    if prog_id:
        indvd_qs = indvd_qs.filter(anix_id__evnt_id__prog_id=prog_id)
        sampd_qs = sampd_qs.filter(samp_id__anix_id__evnt_id__prog_id=prog_id)
    if year:
        indvd_qs = indvd_qs.filter(anix_id__indv_id__indv_year=year)
        sampd_qs = sampd_qs.filter(samp_id__anix_id__grp_id__grp_year=year)

    row_count = 3
    for indvd in indvd_qs:
        indv_id = indvd.anix_id.indv_id
        evnt_id = indvd.anix_id.evnt_id
        ws["A" + str(row_count)].value = evnt_id.__str__()
        ws["B" + str(row_count)].value = indvd.detail_date
        ws["C" + str(row_count)].value = indvd.det_val
        if anidc_name == "Vial":
            ws["D" + str(row_count)].value = indv_id.individual_detail(anidc_name="Box Location", evnt_id=evnt_id,
                                                                         before_date=indvd.detail_date)
            ws["E" + str(row_count)].value = indv_id.individual_detail(anidc_name="Box", evnt_id=evnt_id,
                                                                         before_date=indvd.detail_date)
        ws["F" + str(row_count)].value = indv_id.pit_tag
        ws["H" + str(row_count)].value = indv_id.stok_id.__str__()
        ws["I" + str(row_count)].value = "{}, {}".format(indv_id.indv_year, indv_id.coll_id.__str__())
        ws['J' + str(row_count)].value = indv_id.individual_detail(anidc_name="Gender", evnt_id=evnt_id,
                                                                         before_date=indvd.detail_date)
        ws['K' + str(row_count)].value = indv_id.individual_detail(anidc_name="Length", evnt_id=evnt_id,
                                                                         before_date=indvd.detail_date)
        ws['L' + str(row_count)].value = indv_id.individual_detail(anidc_name="Weight", evnt_id=evnt_id,
                                                                         before_date=indvd.detail_date)
        ws["M" + str(row_count)].value = indvd.comments
        ws["N" + str(row_count)].value = request.build_absolute_uri(reverse("bio_diversity:details_evnt", args=[evnt_id.id]))
        row_count += 1

    for sampd in sampd_qs:
        samp_id = sampd.samp_id
        grp_id = samp_id.anix_id.grp_id
        evnt_id = samp_id.anix_id.evnt_id
        ws["A" + str(row_count)].value = evnt_id.__str__()
        ws["B" + str(row_count)].value = sampd.detail_date
        ws["C" + str(row_count)].value = sampd.det_val
        ws["D" + str(row_count)].value = samp_id.sample_detail(anidc_name="Box Location", evnt_id=evnt_id,
                                                                   before_date=indvd.detail_date)
        ws["E" + str(row_count)].value = samp_id.sample_detail(anidc_name="Box", evnt_id=evnt_id,
                                                                   before_date=indvd.detail_date)
        ws["G" + str(row_count)].value = samp_id.samp_num
        ws["H" + str(row_count)].value = grp_id.stok_id.__str__()
        ws["I" + str(row_count)].value = "{}, {}".format(grp_id.grp_year, grp_id.coll_id.__str__())
        ws['J' + str(row_count)].value = samp_id.sample_detail(anidc_name="Gender", evnt_id=evnt_id,
                                                               before_date=sampd.detail_date)
        ws['K' + str(row_count)].value = samp_id.sample_detail(anidc_name="Length", evnt_id=evnt_id,
                                                               before_date=sampd.detail_date)
        ws['L' + str(row_count)].value = samp_id.sample_detail(anidc_name="Weight", evnt_id=evnt_id,
                                                               before_date=sampd.detail_date)
        ws["M" + str(row_count)].value = sampd.comments
        ws["N" + str(row_count)].value = request.build_absolute_uri(reverse("bio_diversity:details_evnt", args=[evnt_id.id]))
        row_count += 1

    report.save_wb()
    return
